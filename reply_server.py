from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form, Body, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi import Response, Cookie
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import List, Tuple, Optional, Dict, Any
from pathlib import Path
from urllib.parse import unquote
import hashlib
import secrets
import time
import json
import os
import re
import uvicorn
import pandas as pd
import io
import asyncio
import sqlite3
from collections import defaultdict

import cookie_manager
from db_manager import db_manager
from file_log_collector import setup_file_logging, get_file_log_collector
from ai_reply_engine import ai_reply_engine
from utils.qr_login import qr_login_manager
from utils.xianyu_utils import trans_cookies
from utils.image_utils import image_manager

from loguru import logger

# 刮刮乐远程控制路由
try:
    from api_captcha_remote import router as captcha_router
    CAPTCHA_ROUTER_AVAILABLE = True
except ImportError:
    logger.warning("⚠️ api_captcha_remote 未找到，刮刮乐远程控制功能不可用")
    CAPTCHA_ROUTER_AVAILABLE = False

# 关键字文件路径
KEYWORDS_FILE = Path(__file__).parent / "回复关键字.txt"

# 简单的用户认证配置
ADMIN_USERNAME = "admin"

# Session Cookie 配置（安全优先）
SESSION_COOKIE_NAME = "session"
SESSION_EXPIRE_SECONDS = 24 * 60 * 60


def _get_db_path() -> str:
    return os.getenv('DB_PATH', 'data/xianyu_data.db')


def _init_sessions_table_if_needed():
    conn = sqlite3.connect(_get_db_path(), check_same_thread=False)
    try:
        cursor = conn.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS sessions (
            session_id TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL,
            username TEXT NOT NULL,
            is_admin INTEGER DEFAULT 0,
            expires_at INTEGER NOT NULL,
            created_at INTEGER NOT NULL
        )
        ''')
        conn.commit()
    finally:
        conn.close()


def _create_session(user: Dict[str, Any]) -> str:
    _init_sessions_table_if_needed()
    session_id = secrets.token_urlsafe(32)
    now = int(time.time())
    expires_at = now + SESSION_EXPIRE_SECONDS

    conn = sqlite3.connect(_get_db_path(), check_same_thread=False)
    try:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO sessions (session_id, user_id, username, is_admin, expires_at, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (
                session_id,
                int(user['id']),
                str(user['username']),
                1 if user.get('is_admin') else 0,
                int(expires_at),
                int(now),
            ),
        )
        conn.commit()
        return session_id
    finally:
        conn.close()


def _delete_session(session_id: str) -> None:
    _init_sessions_table_if_needed()
    conn = sqlite3.connect(_get_db_path(), check_same_thread=False)
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM sessions WHERE session_id = ?", (session_id,))
        conn.commit()
    finally:
        conn.close()


def _get_session(session_id: str) -> Optional[Dict[str, Any]]:
    if not session_id:
        return None

    _init_sessions_table_if_needed()
    conn = sqlite3.connect(_get_db_path(), check_same_thread=False)
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT session_id, user_id, username, is_admin, expires_at FROM sessions WHERE session_id = ?",
            (session_id,),
        )
        row = cursor.fetchone()
        if not row:
            return None

        now = int(time.time())
        expires_at = int(row[4])
        if expires_at <= now:
            cursor.execute("DELETE FROM sessions WHERE session_id = ?", (session_id,))
            conn.commit()
            return None

        return {
            'session_id': row[0],
            'user_id': int(row[1]),
            'username': row[2],
            'is_admin': bool(row[3]),
            'timestamp': float(now),
        }
    finally:
        conn.close()


def get_current_user_from_session_cookie(session: Optional[str] = Cookie(default=None, alias=SESSION_COOKIE_NAME)) -> Optional[Dict[str, Any]]:
    return _get_session(session)


# ==================== Cookie Session Auth (Phase 1) ====================
# 已切换为 HttpOnly Cookie Session；不再支持 Bearer token。

# 扫码登录检查锁 - 防止并发处理同一个session
qr_check_locks = defaultdict(lambda: asyncio.Lock())
qr_check_processed = {}  # 记录已处理的session: {session_id: {'processed': bool, 'timestamp': float}}

# 账号密码登录会话管理
password_login_sessions = {}  # {session_id: {'account_id': str, 'account': str, 'password': str, 'show_browser': bool, 'status': str, 'verification_url': str, 'qr_code_url': str, 'slider_instance': object, 'task': asyncio.Task, 'timestamp': float}}
password_login_locks = defaultdict(lambda: asyncio.Lock())

# 不再需要单独的密码初始化，由数据库初始化时处理


def cleanup_qr_check_records():
    """清理过期的扫码检查记录"""
    current_time = time.time()
    expired_sessions = []

    for session_id, record in qr_check_processed.items():
        # 清理超过1小时的记录
        if current_time - record['timestamp'] > 3600:
            expired_sessions.append(session_id)

    for session_id in expired_sessions:
        if session_id in qr_check_processed:
            del qr_check_processed[session_id]
        if session_id in qr_check_locks:
            del qr_check_locks[session_id]


def load_keywords() -> List[Tuple[str, str]]:
    """读取关键字→回复映射表

    文件格式支持：
        关键字<空格/制表符/冒号>回复内容
    忽略空行和以 # 开头的注释行
    """
    mapping: List[Tuple[str, str]] = []
    if not KEYWORDS_FILE.exists():
        return mapping

    with KEYWORDS_FILE.open('r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            # 尝试用\t、空格、冒号分隔
            if '\t' in line:
                key, reply = line.split('\t', 1)
            elif ' ' in line:
                key, reply = line.split(' ', 1)
            elif ':' in line:
                key, reply = line.split(':', 1)
            else:
                # 无法解析的行，跳过
                continue
            mapping.append((key.strip(), reply.strip()))
    return mapping


KEYWORDS_MAPPING = load_keywords()


# 认证相关模型
class LoginRequest(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    email: Optional[str] = None
    verification_code: Optional[str] = None


class LoginResponse(BaseModel):
    success: bool
    token: Optional[str] = None
    message: str
    user_id: Optional[int] = None
    username: Optional[str] = None
    is_admin: Optional[bool] = None


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class RegisterRequest(BaseModel):
    username: str
    email: str
    password: str
    verification_code: str


class RegisterResponse(BaseModel):
    success: bool
    message: str


class SendCodeRequest(BaseModel):
    email: str
    session_id: Optional[str] = None
    type: Optional[str] = 'register'  # 'register' 或 'login'


class SendCodeResponse(BaseModel):
    success: bool
    message: str


class CaptchaRequest(BaseModel):
    session_id: str


class CaptchaResponse(BaseModel):
    success: bool
    captcha_image: str
    session_id: str
    message: str


class VerifyCaptchaRequest(BaseModel):
    session_id: str
    captcha_code: str


class VerifyCaptchaResponse(BaseModel):
    success: bool
    message: str


def verify_token() -> Optional[Dict[str, Any]]:
    """已移除：Bearer token 认证不再支持"""
    return None


def verify_session(user_info: Optional[Dict[str, Any]] = Depends(get_current_user_from_session_cookie)) -> Optional[Dict[str, Any]]:
    """验证 HttpOnly Cookie Session 并返回用户信息"""
    return user_info


def verify_admin_token() -> Dict[str, Any]:
    """已移除：Bearer token 管理员认证不再支持"""
    raise HTTPException(status_code=410, detail="Bearer token 认证已移除")


def require_auth(user_info: Optional[Dict[str, Any]] = Depends(verify_session)):
    """需要认证的依赖，返回用户信息（Cookie Session）"""
    if not user_info:
        raise HTTPException(status_code=401, detail="未授权访问")
    return user_info


def get_current_user(user_info: Dict[str, Any] = Depends(require_auth)) -> Dict[str, Any]:
    """获取当前登录用户信息"""
    return user_info


def get_current_user_optional(user_info: Optional[Dict[str, Any]] = Depends(verify_session)) -> Optional[Dict[str, Any]]:
    """获取当前用户信息（可选，不强制要求登录；Cookie Session）"""
    return user_info


def get_user_log_prefix(user_info: Dict[str, Any] = None) -> str:
    """获取用户日志前缀"""
    if user_info:
        return f"【{user_info['username']}#{user_info['user_id']}】"
    return "【系统】"


def require_admin(current_user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    """要求管理员权限"""
    if not current_user.get('is_admin', False):
        raise HTTPException(status_code=403, detail="需要管理员权限")
    return current_user


def log_with_user(level: str, message: str, user_info: Dict[str, Any] = None):
    """带用户信息的日志记录"""
    prefix = get_user_log_prefix(user_info)
    full_message = f"{prefix} {message}"

    if level.lower() == 'info':
        logger.info(full_message)
    elif level.lower() == 'error':
        logger.error(full_message)
    elif level.lower() == 'warning':
        logger.warning(full_message)
    elif level.lower() == 'debug':
        logger.debug(full_message)
    else:
        logger.info(full_message)


def match_reply(cookie_id: str, message: str) -> Optional[str]:
    """根据 cookie_id 及消息内容匹配回复
    只有启用的账号才会匹配关键字回复
    """
    mgr = cookie_manager.manager
    if mgr is None:
        return None

    # 检查账号是否启用
    if not mgr.get_cookie_status(cookie_id):
        return None  # 禁用的账号不参与自动回复

    # 优先账号级关键字
    if mgr.get_keywords(cookie_id):
        for k, r in mgr.get_keywords(cookie_id):
            if k in message:
                return r

    # 全局关键字
    for k, r in KEYWORDS_MAPPING:
        if k in message:
            return r
    return None


class RequestModel(BaseModel):
    cookie_id: str
    msg_time: str
    user_url: str
    send_user_id: str
    send_user_name: str
    item_id: str
    send_message: str
    chat_id: str


class ResponseData(BaseModel):
    send_msg: str


class ResponseModel(BaseModel):
    code: int
    data: ResponseData


app = FastAPI(
    title="Xianyu Auto Reply API",
    version="1.0.0",
    description="闲鱼自动回复系统API",
    docs_url="/docs",
    redoc_url="/redoc"
)

# 添加 CORS 中间件支持前端跨域请求
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:3001",
        "http://localhost:3002",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:3001",
        "http://127.0.0.1:3002",
    ],  # 允许的前端开发服务器地址
    allow_credentials=True,  # 允许携带凭证
    allow_methods=["*"],  # 允许所有HTTP方法
    allow_headers=["*"],  # 允许所有请求头
)

# 注册刮刮乐远程控制路由
if CAPTCHA_ROUTER_AVAILABLE:
    app.include_router(captcha_router)
    logger.info("✅ 已注册刮刮乐远程控制路由: /api/captcha")
else:
    logger.warning("⚠️ 刮刮乐远程控制路由未注册")

# 初始化文件日志收集器
setup_file_logging()

# 添加一条测试日志
from loguru import logger
logger.info("Web服务器启动，文件日志收集器已初始化")

# 添加请求日志中间件
@app.middleware("http")
async def log_requests(request, call_next):
    start_time = time.time()

    logger.info(f"🌐 API请求: {request.method} {request.url.path}")

    response = await call_next(request)

    process_time = time.time() - start_time
    logger.info(f"✅ API响应: {request.method} {request.url.path} - {response.status_code} ({process_time:.3f}s)")

    return response

# 提供前端静态文件
import os
static_dir = os.path.join(os.path.dirname(__file__), 'static')
if not os.path.exists(static_dir):
    os.makedirs(static_dir, exist_ok=True)

# 挂载静态文件目录
app.mount('/static', StaticFiles(directory=static_dir), name='static')

# 挂载 /assets 路径，指向 static/assets 目录
# 这样访问 /assets/xxx.js 时会正确映射到 static_dir/assets/xxx.js
assets_dir = os.path.join(static_dir, 'assets')
if os.path.exists(assets_dir):
    app.mount('/assets', StaticFiles(directory=assets_dir), name='assets')

# 确保图片上传目录存在
uploads_dir = os.path.join(static_dir, 'uploads', 'images')
if not os.path.exists(uploads_dir):
    os.makedirs(uploads_dir, exist_ok=True)
    logger.info(f"创建图片上传目录: {uploads_dir}")

# 健康检查端点
@app.get('/health')
async def health_check():
    """健康检查端点，用于Docker健康检查和负载均衡器"""
    try:
        # 检查Cookie管理器状态
        manager_status = "ok" if cookie_manager.manager is not None else "error"

        # 检查数据库连接
        from db_manager import db_manager
        try:
            db_manager.get_all_cookies()
            db_status = "ok"
        except Exception:
            db_status = "error"

        # 获取系统状态
        import psutil
        cpu_percent = psutil.cpu_percent(interval=1)
        memory_info = psutil.virtual_memory()

        status = {
            "status": "healthy" if manager_status == "ok" and db_status == "ok" else "unhealthy",
            "timestamp": time.time(),
            "services": {
                "cookie_manager": manager_status,
                "database": db_status
            },
            "system": {
                "cpu_percent": cpu_percent,
                "memory_percent": memory_info.percent,
                "memory_available": memory_info.available
            }
        }

        if status["status"] == "unhealthy":
            raise HTTPException(status_code=503, detail=status)

        return status

    except Exception as e:
        return {
            "status": "unhealthy",
            "timestamp": time.time(),
            "error": str(e)
        }


# 服务 React 前端 SPA - 所有前端路由都返回 index.html
async def serve_frontend():
    """服务 React 前端 SPA"""
    index_path = os.path.join(static_dir, 'index.html')
    if os.path.exists(index_path):
        with open(index_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Frontend not found. Please build the frontend first.</h3>')

@app.get('/', response_class=HTMLResponse)
async def root():
    return await serve_frontend()


# 登录页面路由 - 重定向到 React 前端
@app.get('/login.html', response_class=HTMLResponse)
async def login_page():
    return await serve_frontend()

@app.get('/login', response_class=HTMLResponse)
async def login_route():
    return await serve_frontend()


# 初始化页面路由
@app.get('/init', response_class=HTMLResponse)
async def init_route():
    return await serve_frontend()


# 注册页面路由
@app.get('/register.html', response_class=HTMLResponse)
async def register_page():
    # 检查注册是否开启
    from db_manager import db_manager
    registration_enabled = db_manager.get_system_setting('registration_enabled')
    if registration_enabled != 'true':
        return HTMLResponse('''
        <!DOCTYPE html>
        <html>
        <head>
            <title>注册已关闭</title>
            <meta charset="utf-8">
            <style>
                body { font-family: Arial, sans-serif; text-align: center; padding: 50px; }
                .message { color: #666; font-size: 18px; }
                .back-link { margin-top: 20px; }
                .back-link a { color: #007bff; text-decoration: none; }
            </style>
        </head>
        <body>
            <h2>🚫 注册功能已关闭</h2>
            <p class="message">系统管理员已关闭用户注册功能</p>
            <div class="back-link">
                <a href="/">← 返回首页</a>
            </div>
        </body>
        </html>
        ''', status_code=403)

    return await serve_frontend()

# 注意：不要在这里定义 /admin 或 /admin/{path} 路由
# 因为后端有 /admin/users, /admin/logs 等 API 路由
# 前端 SPA 通过根路由 / 加载，由 React Router 处理客户端路由
# 文件末尾的 catch-all 路由会处理前端页面的直接访问



# 登录接口
@app.post('/login')
async def login(request: LoginRequest):
    from db_manager import db_manager

    # 判断登录方式
    if request.username and request.password:
        # 用户名/密码登录
        logger.info(f"【{request.username}】尝试用户名登录")

        # 统一使用用户表验证（包括admin用户）
        if db_manager.verify_user_password(request.username, request.password):
            user = db_manager.get_user_by_username(request.username)
            if user:
                # 创建 Cookie Session
                session_id = _create_session({
                    'id': user['id'],
                    'username': user['username'],
                    'is_admin': bool(user.get('is_admin', False)) or user['username'] == ADMIN_USERNAME,
                })

                # 区分管理员和普通用户的日志
                if user['username'] == ADMIN_USERNAME:
                    logger.info(f"【{user['username']}#{user['id']}】登录成功（管理员）")
                else:
                    logger.info(f"【{user['username']}#{user['id']}】登录成功")

                resp = JSONResponse(content=LoginResponse(
                    success=True,
                    token=None,
                    message="登录成功",
                    user_id=user['id'],
                    username=user['username'],
                    is_admin=bool(user.get('is_admin', False)) or user['username'] == ADMIN_USERNAME
                ).model_dump())
                resp.set_cookie(
                    key=SESSION_COOKIE_NAME,
                    value=session_id,
                    httponly=True,
                    samesite='lax',
                    secure=False,
                    max_age=SESSION_EXPIRE_SECONDS,
                    path='/',
                )
                return resp

        logger.warning(f"【{request.username}】登录失败：用户名或密码错误")
        return LoginResponse(
            success=False,
            message="用户名或密码错误"
        )

    elif request.email and request.password:
        # 邮箱/密码登录
        logger.info(f"【{request.email}】尝试邮箱密码登录")

        user = db_manager.get_user_by_email(request.email)
        if user and db_manager.verify_user_password(user['username'], request.password):
            # 创建 Cookie Session
            session_id = _create_session({
                'id': user['id'],
                'username': user['username'],
                'is_admin': bool(user.get('is_admin', False)) or user['username'] == ADMIN_USERNAME,
            })

            logger.info(f"【{user['username']}#{user['id']}】邮箱登录成功")

            resp = JSONResponse(content=LoginResponse(
                success=True,
                token=None,
                message="登录成功",
                user_id=user['id'],
                username=user['username'],
                is_admin=bool(user.get('is_admin', False)) or user['username'] == ADMIN_USERNAME
            ).model_dump())
            resp.set_cookie(
                key=SESSION_COOKIE_NAME,
                value=session_id,
                httponly=True,
                samesite='lax',
                secure=False,
                max_age=SESSION_EXPIRE_SECONDS,
                path='/',
            )
            return resp

        logger.warning(f"【{request.email}】邮箱登录失败：邮箱或密码错误")
        return LoginResponse(
            success=False,
            message="邮箱或密码错误"
        )

    elif request.email and request.verification_code:
        # 邮箱/验证码登录
        logger.info(f"【{request.email}】尝试邮箱验证码登录")

        # 验证邮箱验证码
        if not db_manager.verify_email_code(request.email, request.verification_code, 'login'):
            logger.warning(f"【{request.email}】验证码登录失败：验证码错误或已过期")
            return LoginResponse(
                success=False,
                message="验证码错误或已过期"
            )

        # 获取用户信息
        user = db_manager.get_user_by_email(request.email)
        if not user:
            logger.warning(f"【{request.email}】验证码登录失败：用户不存在")
            return LoginResponse(
                success=False,
                message="用户不存在"
            )

        # 创建 Cookie Session
        session_id = _create_session({
            'id': user['id'],
            'username': user['username'],
            'is_admin': bool(user.get('is_admin', False)) or user['username'] == ADMIN_USERNAME,
        })

        logger.info(f"【{user['username']}#{user['id']}】验证码登录成功")

        resp = JSONResponse(content=LoginResponse(
            success=True,
            token=None,
            message="登录成功",
            user_id=user['id'],
            username=user['username'],
            is_admin=bool(user.get('is_admin', False)) or user['username'] == ADMIN_USERNAME
        ).model_dump())
        resp.set_cookie(
            key=SESSION_COOKIE_NAME,
            value=session_id,
            httponly=True,
            samesite='lax',
            secure=False,
            max_age=SESSION_EXPIRE_SECONDS,
            path='/',
        )
        return resp

    else:
        return LoginResponse(
            success=False,
            message="请提供有效的登录信息"
        )


# 验证token接口
@app.get('/verify')
async def verify(user_info: Optional[Dict[str, Any]] = Depends(verify_session)):
    # 如果系统尚未初始化（没有 admin 用户），前端需要显示初始化引导
    from db_manager import db_manager
    initialized = db_manager.is_system_initialized()

    if user_info:
        return {
            "authenticated": True,
            "user_id": user_info['user_id'],
            "username": user_info['username'],
            "is_admin": bool(user_info.get('is_admin', False)),
            "initialized": initialized,
        }

    return {
        "authenticated": False,
        "initialized": initialized,
    }


# 登出接口
@app.post('/logout')
async def logout(response: Response, session: Optional[str] = Cookie(default=None, alias=SESSION_COOKIE_NAME)):
    if session:
        _delete_session(session)
    response = JSONResponse(content={"message": "已登出"})
    response.delete_cookie(key=SESSION_COOKIE_NAME, path='/')
    return response


# 修改管理员密码接口
@app.post('/change-admin-password')
async def change_admin_password(request: ChangePasswordRequest, admin_user: Dict[str, Any] = Depends(require_admin)):
    from db_manager import db_manager

    try:
        # 验证当前密码（使用用户表验证）
        if not db_manager.verify_user_password(ADMIN_USERNAME, request.current_password):
            return {"success": False, "message": "当前密码错误"}

        # 更新密码（使用用户表更新）
        success = db_manager.update_user_password(ADMIN_USERNAME, request.new_password)

        if success:
            logger.info(f"【admin#{admin_user['user_id']}】管理员密码修改成功")
            return {"success": True, "message": "密码修改成功"}
        else:
            return {"success": False, "message": "密码修改失败"}

    except Exception as e:
        logger.error(f"修改管理员密码异常: {e}")
        return {"success": False, "message": "系统错误"}


# 普通用户修改密码接口
@app.post('/change-password')
async def change_user_password(request: ChangePasswordRequest, current_user: Dict[str, Any] = Depends(get_current_user)):
    from db_manager import db_manager

    try:
        username = current_user.get('username')
        user_id = current_user.get('user_id')
        
        if not username:
            return {"success": False, "message": "无法获取用户信息"}

        # 验证当前密码
        if not db_manager.verify_user_password(username, request.current_password):
            return {"success": False, "message": "当前密码错误"}

        # 更新密码
        success = db_manager.update_user_password(username, request.new_password)

        if success:
            logger.info(f"【{username}#{user_id}】用户密码修改成功")
            return {"success": True, "message": "密码修改成功"}
        else:
            return {"success": False, "message": "密码修改失败"}

    except Exception as e:
        logger.error(f"修改用户密码异常: {e}")
        return {"success": False, "message": "系统错误"}


# 检查是否使用默认密码
# 安全原因：不再支持“默认口令检查”接口。默认口令会导致高危风险，且该接口曾在日志中泄露口令。
# ���需管理员初始化/重置，请通过受控的初始化流程或管理员修改密码接口完成。
@app.get('/api/check-default-password')
async def check_default_password(current_user: Dict[str, Any] = Depends(get_current_user)):
    raise HTTPException(status_code=404, detail="接口已移除")


# 生成图形验证码接口
@app.post('/generate-captcha')
async def generate_captcha(request: CaptchaRequest):
    from db_manager import db_manager

    try:
        # 生成图形验证码
        captcha_text, captcha_image = db_manager.generate_captcha()

        if not captcha_image:
            return CaptchaResponse(
                success=False,
                captcha_image="",
                session_id=request.session_id,
                message="图形验证码生成失败"
            )

        # 保存验证码到数据库
        if db_manager.save_captcha(request.session_id, captcha_text):
            return CaptchaResponse(
                success=True,
                captcha_image=captcha_image,
                session_id=request.session_id,
                message="图形验证码生成成功"
            )
        else:
            return CaptchaResponse(
                success=False,
                captcha_image="",
                session_id=request.session_id,
                message="图形验证码保存失败"
            )

    except Exception as e:
        logger.error(f"生成图形验证码失败: {e}")
        return CaptchaResponse(
            success=False,
            captcha_image="",
            session_id=request.session_id,
            message="图形验证码生成失败"
        )


# 验证图形验证码接口
@app.post('/verify-captcha')
async def verify_captcha(request: VerifyCaptchaRequest):
    from db_manager import db_manager

    try:
        if db_manager.verify_captcha(request.session_id, request.captcha_code):
            return VerifyCaptchaResponse(
                success=True,
                message="图形验证码验证成功"
            )
        else:
            return VerifyCaptchaResponse(
                success=False,
                message="图形验证码错误或已过期"
            )

    except Exception as e:
        logger.error(f"验证图形验证码失败: {e}")
        return VerifyCaptchaResponse(
            success=False,
            message="图形验证码验证失败"
        )


# ==================== 极验滑动验证码 ====================

# 极验验证状态存储: {challenge: {"status": int, "expires_at": float}}
geetest_status_store: dict = {}


def cleanup_expired_geetest_status():
    """清理过期的极验验证状态"""
    current_time = time.time()
    expired_keys = [k for k, v in geetest_status_store.items() if v["expires_at"] < current_time]
    for k in expired_keys:
        del geetest_status_store[k]


def set_geetest_status(challenge: str, status: int):
    """设置极验验证状态"""
    cleanup_expired_geetest_status()
    geetest_status_store[challenge] = {
        "status": status,
        "expires_at": time.time() + 300  # 5分钟有效
    }


def get_geetest_status(challenge: str) -> int:
    """获取极验验证状态，返回0表示未验证或已过期"""
    cleanup_expired_geetest_status()
    stored = geetest_status_store.get(challenge)
    if stored and stored["expires_at"] > time.time():
        return stored["status"]
    return 0


class GeetestRegisterResponse(BaseModel):
    """极验验证码初始化响应"""
    success: bool
    code: int = 200
    message: str = ""
    data: Optional[dict] = None


class GeetestValidateRequest(BaseModel):
    """极验二次验证请求"""
    challenge: str
    validate_str: str = Field(..., alias='validate')
    seccode: str

    model_config = {'populate_by_name': True}


class GeetestValidateResponse(BaseModel):
    """极验二次验证响应"""
    success: bool
    code: int = 200
    message: str = ""


@app.get('/geetest/register', response_model=GeetestRegisterResponse)
async def geetest_register():
    """
    获取极验验证码初始化参数
    
    前端调用此接口获取gt、challenge等参数，用于初始化验证码组件
    """
    try:
        from utils.geetest import GeetestLib
        
        gt_lib = GeetestLib()
        result = await gt_lib.register()
        
        data = result.to_dict()
        logger.info(f"极验初始化结果: status={result.status}, data={data}")
        
        # 记录初始状态
        challenge = data.get("challenge", "")
        if challenge:
            set_geetest_status(challenge, 0)
        
        return GeetestRegisterResponse(
            success=True,
            code=200,
            message="获取成功" if result.status == 1 else "宕机模式",
            data=data
        )
            
    except Exception as e:
        logger.error(f"极验初始化失败: {e}")
        # 返回本地初始化结果
        try:
            from utils.geetest import GeetestLib
            gt_lib = GeetestLib()
            result = gt_lib.local_init()
            data = result.to_dict()
            
            # 记录初始状态
            challenge = data.get("challenge", "")
            if challenge:
                set_geetest_status(challenge, 0)
            
            return GeetestRegisterResponse(
                success=True,
                code=200,
                message="本地初始化",
                data=data
            )
        except Exception as e2:
            logger.error(f"极验本地初始化也失败: {e2}")
            return GeetestRegisterResponse(
                success=False,
                code=500,
                message="验证码服务异常"
            )


@app.post('/geetest/validate', response_model=GeetestValidateResponse)
async def geetest_validate(request: GeetestValidateRequest):
    """
    极验二次验证
    
    用户完成滑动验证后，前端调用此接口进行二次验证
    """
    try:
        # 检查是否已经验证过
        if get_geetest_status(request.challenge) == 1:
            return GeetestValidateResponse(
                success=True,
                code=200,
                message="验证通过"
            )
        
        from utils.geetest import GeetestLib
        
        gt_lib = GeetestLib()
        
        # 判断是正常模式还是宕机模式
        # 通过challenge长度判断：正常模式challenge是32位MD5，宕机模式是UUID
        is_normal_mode = len(request.challenge) == 32
        
        if is_normal_mode:
            result = await gt_lib.success_validate(
                request.challenge,
                request.validate_str,
                request.seccode
            )
        else:
            result = gt_lib.fail_validate(
                request.challenge,
                request.validate_str,
                request.seccode
            )
        
        if result.status == 1:
            # 记录验证通过状态
            set_geetest_status(request.challenge, 1)
            
            return GeetestValidateResponse(
                success=True,
                code=200,
                message="验证通过"
            )
        else:
            return GeetestValidateResponse(
                success=False,
                code=400,
                message=result.msg or "验证失败"
            )
            
    except Exception as e:
        logger.error(f"极验二次验证失败: {e}")
        return GeetestValidateResponse(
            success=False,
            code=500,
            message="验证服务异常"
        )


# 发送验证码接口（需要先验证图形验证码）
@app.post('/send-verification-code')
async def send_verification_code(request: SendCodeRequest):
    from db_manager import db_manager

    try:
        # 检查是否已验证图形验证码
        # 通过检查数据库中是否存在已验证的图形验证码记录
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            current_time = time.time()

            # 查找最近5分钟内该session_id的验证记录
            # 由于验证成功后验证码会被删除，我们需要另一种方式来跟踪验证状态
            # 这里我们检查该session_id是否在最近验证过（通过检查是否有已删除的记录）

            # 为了简化，我们要求前端在验证图形验证码成功后立即发送邮件验证码
            # 或者我们可以在验证成功后设置一个临时标记
            pass

        # 根据验证码类型进行不同的检查
        if request.type == 'register':
            # 注册验证码：检查邮箱是否已注册
            existing_user = db_manager.get_user_by_email(request.email)
            if existing_user:
                return SendCodeResponse(
                    success=False,
                    message="该邮箱已被注册"
                )
        elif request.type == 'login':
            # 登录验证码：检查邮箱是否存在
            existing_user = db_manager.get_user_by_email(request.email)
            if not existing_user:
                return SendCodeResponse(
                    success=False,
                    message="该邮箱未注册"
                )

        # 生成验证码
        code = db_manager.generate_verification_code()

        # 保存验证码到数据库
        if not db_manager.save_verification_code(request.email, code, request.type):
            return SendCodeResponse(
                success=False,
                message="验证码保存失败，请稍后重试"
            )

        # 发送验证码邮件
        if await db_manager.send_verification_email(request.email, code):
            return SendCodeResponse(
                success=True,
                message="验证码已发送到您的邮箱，请查收"
            )
        else:
            return SendCodeResponse(
                success=False,
                message="验证码发送失败，请检查邮箱地址或稍后重试"
            )

    except Exception as e:
        logger.error(f"发送验证码失败: {e}")
        return SendCodeResponse(
            success=False,
            message="发送验证码失败，请稍后重试"
        )


# 用户注册接口
@app.post('/register')
async def register(request: RegisterRequest):
    from db_manager import db_manager

    # 检查注册是否开启
    registration_enabled = db_manager.get_system_setting('registration_enabled')
    if registration_enabled != 'true':
        logger.warning(f"【{request.username}】注册失败: 注册功能已关闭")
        return RegisterResponse(
            success=False,
            message="注册功能已关闭，请联系管理员"
        )

    try:
        logger.info(f"【{request.username}】尝试注册，邮箱: {request.email}")

        # 验证邮箱验证码
        if not db_manager.verify_email_code(request.email, request.verification_code):
            logger.warning(f"【{request.username}】注册失败: 验证码错误或已过期")
            return RegisterResponse(
                success=False,
                message="验证码错误或已过期"
            )

        # 检查用户名是否已存在
        existing_user = db_manager.get_user_by_username(request.username)
        if existing_user:
            logger.warning(f"【{request.username}】注册失败: 用户名已存在")
            return RegisterResponse(
                success=False,
                message="用户名已存在"
            )

        # 检查邮箱是否已注册
        existing_email = db_manager.get_user_by_email(request.email)
        if existing_email:
            logger.warning(f"【{request.username}】注册失败: 邮箱已被注册")
            return RegisterResponse(
                success=False,
                message="该邮箱已被注册"
            )

        # 创建用户
        if db_manager.create_user(request.username, request.email, request.password):
            logger.info(f"【{request.username}】注册成功")
            return RegisterResponse(
                success=True,
                message="注册成功，请登录"
            )
        else:
            logger.error(f"【{request.username}】注册失败: 数据库操作失败")
            return RegisterResponse(
                success=False,
                message="注册失败，请稍后重试"
            )

    except Exception as e:
        logger.error(f"【{request.username}】注册异常: {e}")
        return RegisterResponse(
            success=False,
            message="注册失败，请稍后重试"
        )


# ------------------------- 发送消息接口 -------------------------

class SendMessageRequest(BaseModel):
    api_key: str
    cookie_id: str
    chat_id: str
    to_user_id: str
    message: str


class SendMessageResponse(BaseModel):
    success: bool
    message: str


def verify_api_key(api_key: str) -> bool:
    """验证API秘钥

    安全基线：必须显式配置（系统设置/环境变量），不再允许硬编码后备key。
    """
    try:
        from db_manager import db_manager
        qq_secret_key = db_manager.get_system_setting('qq_reply_secret_key')

        if not qq_secret_key:
            logger.error("qq_reply_secret_key 未配置，拒绝请求")
            return False

        return api_key == qq_secret_key
    except Exception as e:
        logger.error(f"验证API秘钥时发生异常: {e}")
        return False


@app.post('/send-message', response_model=SendMessageResponse)
async def send_message_api(request: SendMessageRequest):
    """发送消息API接口（使用秘钥验证）"""
    try:
        # 清理所有参数中的换行符
        def clean_param(param_str):
            """清理参数中的换行符"""
            if isinstance(param_str, str):
                return param_str.replace('\\n', '').replace('\n', '')
            return param_str

        # 清理所有参数
        cleaned_api_key = clean_param(request.api_key)
        cleaned_cookie_id = clean_param(request.cookie_id)
        cleaned_chat_id = clean_param(request.chat_id)
        cleaned_to_user_id = clean_param(request.to_user_id)
        cleaned_message = clean_param(request.message)

        # 验证API秘钥不能为空
        if not cleaned_api_key:
            logger.warning("API秘钥为空")
            return SendMessageResponse(
                success=False,
                message="API秘钥不能为空"
            )

        # 验证API秘钥
        if not verify_api_key(cleaned_api_key):
            logger.warning(f"API秘钥验证失败: {cleaned_api_key}")
            return SendMessageResponse(
                success=False,
                message="API秘钥验证失败"
            )

        # 验证必需参数不能为空
        required_params = {
            'cookie_id': cleaned_cookie_id,
            'chat_id': cleaned_chat_id,
            'to_user_id': cleaned_to_user_id,
            'message': cleaned_message
        }

        for param_name, param_value in required_params.items():
            if not param_value:
                logger.warning(f"必需参数 {param_name} 为空")
                return SendMessageResponse(
                    success=False,
                    message=f"参数 {param_name} 不能为空"
                )

        # 直接获取XianyuLive实例，跳过cookie_manager检查
        from XianyuAutoAsync import XianyuLive
        live_instance = XianyuLive.get_instance(cleaned_cookie_id)

        if not live_instance:
            logger.warning(f"账号实例不存在或未连接: {cleaned_cookie_id}")
            return SendMessageResponse(
                success=False,
                message="账号实例不存在或未连接，请检查账号状态"
            )

        # 检查WebSocket连接状态
        if not live_instance.ws or live_instance.ws.closed:
            logger.warning(f"账号WebSocket连接已断开: {cleaned_cookie_id}")
            return SendMessageResponse(
                success=False,
                message="账号WebSocket连接已断开，请等待重连"
            )

        # 发送消息（使用清理后的所有参数）
        await live_instance.send_msg(
            live_instance.ws,
            cleaned_chat_id,
            cleaned_to_user_id,
            cleaned_message
        )

        logger.info(f"API成功发送消息: {cleaned_cookie_id} -> {cleaned_to_user_id}, 内容: {cleaned_message[:50]}{'...' if len(cleaned_message) > 50 else ''}")

        return SendMessageResponse(
            success=True,
            message="消息发送成功"
        )

    except Exception as e:
        # 使用清理后的参数记录日志
        cookie_id_for_log = clean_param(request.cookie_id) if 'clean_param' in locals() else request.cookie_id
        to_user_id_for_log = clean_param(request.to_user_id) if 'clean_param' in locals() else request.to_user_id
        logger.error(f"API发送消息异常: {cookie_id_for_log} -> {to_user_id_for_log}, 错误: {str(e)}")
        return SendMessageResponse(
            success=False,
            message=f"发送消息失败: {str(e)}"
        )


@app.post("/xianyu/reply", response_model=ResponseModel)
async def xianyu_reply(req: RequestModel):
    msg_template = match_reply(req.cookie_id, req.send_message)
    is_default_reply = False

    if not msg_template:
        # 从数据库获取默认回复
        from db_manager import db_manager
        default_reply_settings = db_manager.get_default_reply(req.cookie_id)

        if default_reply_settings and default_reply_settings.get('enabled', False):
            # 检查是否开启了"只回复一次"功能
            if default_reply_settings.get('reply_once', False):
                # 检查是否已经回复过这个chat_id
                if db_manager.has_default_reply_record(req.cookie_id, req.chat_id):
                    raise HTTPException(status_code=404, detail="该对话已使用默认回复，不再重复回复")

            msg_template = default_reply_settings.get('reply_content', '')
            is_default_reply = True

        # 如果数据库中没有设置或为空，返回错误
        if not msg_template:
            raise HTTPException(status_code=404, detail="未找到匹配的回复规则且未设置默认回复")

    # 按占位符格式化
    try:
        send_msg = msg_template.format(
            send_user_id=req.send_user_id,
            send_user_name=req.send_user_name,
            send_message=req.send_message,
        )
    except Exception:
        # 如果格式化失败，返回原始内容
        send_msg = msg_template

    # 如果是默认回复且开启了"只回复一次"，记录回复记录
    if is_default_reply:
        from db_manager import db_manager
        default_reply_settings = db_manager.get_default_reply(req.cookie_id)
        if default_reply_settings and default_reply_settings.get('reply_once', False):
            db_manager.add_default_reply_record(req.cookie_id, req.chat_id)

    return {"code": 200, "data": {"send_msg": send_msg}}

# ------------------------- 账号 / 关键字管理接口 -------------------------


class CookieIn(BaseModel):
    id: str
    value: str


class CookieStatusIn(BaseModel):
    enabled: bool


class DefaultReplyIn(BaseModel):
    enabled: bool
    reply_content: Optional[str] = None
    reply_image_url: Optional[str] = None
    reply_once: bool = False


class NotificationChannelIn(BaseModel):
    name: str
    type: str = "qq"
    config: str


class NotificationChannelUpdate(BaseModel):
    name: str
    config: str
    enabled: bool = True


class MessageNotificationIn(BaseModel):
    channel_id: int
    enabled: bool = True


class SystemSettingIn(BaseModel):
    value: str
    description: Optional[str] = None


class SystemSettingCreateIn(BaseModel):
    key: str
    value: str
    description: Optional[str] = None





@app.get("/cookies")
def list_cookies(current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        return []

    # 获取当前用户的cookies
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)
    return list(user_cookies.keys())


@app.get("/cookies/details")
def get_cookies_details(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取所有账号的非敏感信息

    安全基线：禁止返回闲鱼Cookie明文、账号密码明文。
    """
    if cookie_manager.manager is None:
        return []

    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    result = []
    for cookie_id in user_cookies.keys():
        cookie_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
        auto_confirm = db_manager.get_auto_confirm(cookie_id)
        cookie_details = db_manager.get_cookie_details(cookie_id)
        remark = cookie_details.get('remark', '') if cookie_details else ''

        result.append({
            'id': cookie_id,
            'has_cookie': True,
            'enabled': cookie_enabled,
            'auto_confirm': auto_confirm,
            'remark': remark,
            'pause_duration': cookie_details.get('pause_duration', 10) if cookie_details else 10
        })

    return result


@app.post("/cookies")
def add_cookie(item: CookieIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 添加cookie时绑定到当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager

        log_with_user('info', f"尝试添加Cookie: {item.id}, 当前用户ID: {user_id}, 用户名: {current_user.get('username', 'unknown')}", current_user)

        # 检查cookie是否已存在且属于其他用户
        existing_cookies = db_manager.get_all_cookies()
        if item.id in existing_cookies:
            # 检查是否属于当前用户
            user_cookies = db_manager.get_all_cookies(user_id)
            if item.id not in user_cookies:
                log_with_user('warning', f"Cookie ID冲突: {item.id} 已被其他用户使用", current_user)
                raise HTTPException(status_code=400, detail="该Cookie ID已被其他用户使用")

        # 保存到数据库时指定用户ID
        db_manager.save_cookie(item.id, item.value, user_id)

        # 添加到CookieManager，同时指定用户ID
        cookie_manager.manager.add_cookie(item.id, item.value, user_id=user_id)
        log_with_user('info', f"Cookie添加成功: {item.id}", current_user)
        return {"msg": "success"}
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"添加Cookie失败: {item.id} - {str(e)}", current_user)
        raise HTTPException(status_code=400, detail=str(e))


# ============ 带子路径的 /cookies/{cid}/xxx 路由必须在 /cookies/{cid} 之前定义 ============

class AccountLoginInfoUpdate(BaseModel):
    username: Optional[str] = None
    login_password: Optional[str] = None
    show_browser: Optional[bool] = None


@app.put("/cookies/{cid}/login-info")
def update_cookie_login_info(cid: str, update_data: AccountLoginInfoUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号登录信息（用户名、密码、是否显示浏览器）"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 使用现有的update_cookie_account_info方法更新登录信息
        success = db_manager.update_cookie_account_info(
            cid,
            username=update_data.username,
            password=update_data.login_password,
            show_browser=update_data.show_browser
        )

        if success:
            return {"success": True, "message": "登录信息已更新"}
        else:
            raise HTTPException(status_code=500, detail="更新登录信息失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ 通用的 /cookies/{cid} 路由 ============

@app.put('/cookies/{cid}')
def update_cookie(cid: str, item: CookieIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取旧的 cookie 值，用于判断是否需要重启任务
        old_cookie_details = db_manager.get_cookie_details(cid)
        old_cookie_value = old_cookie_details.get('value') if old_cookie_details else None

        # 使用 update_cookie_account_info 更新（只更新cookie值，不覆盖其他字段）
        success = db_manager.update_cookie_account_info(cid, cookie_value=item.value)
        
        if not success:
            raise HTTPException(status_code=400, detail="更新Cookie失败")
        
        # 只有当 cookie 值真的发生变化时才重启任务
        if item.value != old_cookie_value:
            logger.info(f"Cookie值已变化，重启任务: {cid}")
            cookie_manager.manager.update_cookie(cid, item.value, save_to_db=False)
        else:
            logger.info(f"Cookie值未变化，无需重启任务: {cid}")
        
        return {'msg': 'updated'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class CookieAccountInfo(BaseModel):
    """账号信息更新模型"""
    value: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None
    show_browser: Optional[bool] = None


@app.post("/cookie/{cid}/account-info")
def update_cookie_account_info(cid: str, info: CookieAccountInfo, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号信息（Cookie、用户名、密码、显示浏览器设置）"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取旧的 cookie 值，用于判断是否需要重启任务
        old_cookie_details = db_manager.get_cookie_details(cid)
        old_cookie_value = old_cookie_details.get('value') if old_cookie_details else None
        
        # 更新数据库
        success = db_manager.update_cookie_account_info(
            cid, 
            cookie_value=info.value,
            username=info.username,
            password=info.password,
            show_browser=info.show_browser
        )
        
        if not success:
            raise HTTPException(status_code=400, detail="更新账号信息失败")
        
        # 只有当 cookie 值真的发生变化时才重启任务
        if info.value is not None and info.value != old_cookie_value:
            logger.info(f"Cookie值已变化，重启任务: {cid}")
            cookie_manager.manager.update_cookie(cid, info.value, save_to_db=False)
        else:
            logger.info(f"Cookie值未变化，无需重启任务: {cid}")
        
        return {'msg': 'updated', 'success': True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新账号信息失败: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/cookie/{cid}/details")
def get_cookie_account_details(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号非敏感详情

    安全基线：禁止返回闲鱼Cookie明文与账号密码明文。
    """
    try:
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        details = db_manager.get_cookie_details(cid)
        if not details:
            raise HTTPException(status_code=404, detail="账号不存在")

        return {
            'id': details.get('id'),
            'enabled': cookie_manager.manager.get_cookie_status(cid) if cookie_manager.manager else True,
            'auto_confirm': details.get('auto_confirm', True),
            'remark': details.get('remark', ''),
            'pause_duration': details.get('pause_duration', 10),
            'show_browser': details.get('show_browser', False),
            'username': details.get('username', ''),
            'has_cookie': True,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取账号详情失败: {e}")
        raise HTTPException(status_code=400, detail=str(e))


# ========================= 账号密码登录相关接口 =========================

async def _execute_password_login(session_id: str, account_id: str, account: str, password: str, show_browser: bool, user_id: int, current_user: Dict[str, Any]):
    """后台执行账号密码登录任务"""
    try:
        log_with_user('info', f"开始执行账号密码登录任务: {session_id}, 账号: {account_id}", current_user)
        
        # 导入 XianyuSliderStealth
        from utils.xianyu_slider_stealth import XianyuSliderStealth
        import base64
        import io
        
        # 创建 XianyuSliderStealth 实例
        slider_instance = XianyuSliderStealth(
            user_id=account_id,
            enable_learning=True,
            headless=not show_browser
        )
        
        # 更新会话信息
        password_login_sessions[session_id]['slider_instance'] = slider_instance
        
        # 定义通知回调函数，用于检测到人脸认证时返回验证链接或截图（同步函数）
        def notification_callback(message: str, screenshot_path: str = None, verification_url: str = None, screenshot_path_new: str = None):
            """人脸认证通知回调（同步）
            
            Args:
                message: 通知消息
                screenshot_path: 旧版截图路径（兼容参数）
                verification_url: 验证链接
                screenshot_path_new: 新版截图路径（新参数，优先使用）
            """
            try:
                # 优先使用新的截图路径参数
                actual_screenshot_path = screenshot_path_new if screenshot_path_new else screenshot_path
                
                # 优先使用截图路径，如果没有截图则使用验证链接
                if actual_screenshot_path and os.path.exists(actual_screenshot_path):
                    # 更新会话状态，保存截图路径
                    password_login_sessions[session_id]['status'] = 'verification_required'
                    password_login_sessions[session_id]['screenshot_path'] = actual_screenshot_path
                    password_login_sessions[session_id]['verification_url'] = None
                    password_login_sessions[session_id]['qr_code_url'] = None
                    log_with_user('info', f"人脸认证截图已保存: {session_id}, 路径: {actual_screenshot_path}", current_user)
                    
                    # 发送通知到用户配置的渠道
                    def send_face_verification_notification():
                        """在后台线程中发送人脸验证通知"""
                        try:
                            from XianyuAutoAsync import XianyuLive
                            log_with_user('info', f"开始尝试发送人脸验证通知: {account_id}", current_user)
                            
                            # 尝试获取XianyuLive实例（如果账号已经存在）
                            live_instance = XianyuLive.get_instance(account_id)
                            
                            if live_instance:
                                log_with_user('info', f"找到账号实例，准备发送通知: {account_id}", current_user)
                                # 创建新的事件循环来运行异步通知
                                new_loop = asyncio.new_event_loop()
                                asyncio.set_event_loop(new_loop)
                                try:
                                    new_loop.run_until_complete(
                                        live_instance.send_token_refresh_notification(
                                            error_message=message,
                                            notification_type="face_verification",
                                            verification_url=None,
                                            attachment_path=actual_screenshot_path
                                        )
                                    )
                                    log_with_user('info', f"✅ 已发送人脸验证通知: {account_id}", current_user)
                                except Exception as notify_err:
                                    log_with_user('error', f"发送人脸验证通知失败: {str(notify_err)}", current_user)
                                    import traceback
                                    log_with_user('error', f"通知错误详情: {traceback.format_exc()}", current_user)
                                finally:
                                    new_loop.close()
                            else:
                                # 如果账号实例不存在，记录警告并尝试从数据库获取通知配置
                                log_with_user('warning', f"账号实例不存在: {account_id}，尝试从数据库获取通知配置", current_user)
                                try:
                                    # 尝试从数据库获取通知配置
                                    notifications = db_manager.get_account_notifications(account_id)
                                    if notifications:
                                        log_with_user('info', f"找到 {len(notifications)} 个通知配置，但需要账号实例才能发送", current_user)
                                        log_with_user('warning', f"账号实例不存在，无法发送通知: {account_id}。请确保账号已登录并运行中。", current_user)
                                    else:
                                        log_with_user('warning', f"账号 {account_id} 未配置通知渠道", current_user)
                                except Exception as db_err:
                                    log_with_user('error', f"获取通知配置失败: {str(db_err)}", current_user)
                        except Exception as notify_err:
                            log_with_user('error', f"发送人脸验证通知时出错: {str(notify_err)}", current_user)
                            import traceback
                            log_with_user('error', f"通知错误详情: {traceback.format_exc()}", current_user)
                    
                    # 在后台线程中发送通知，避免阻塞登录流程
                    import threading
                    notification_thread = threading.Thread(target=send_face_verification_notification)
                    notification_thread.daemon = True
                    notification_thread.start()
                    log_with_user('info', f"已启动人脸验证通知发送线程: {account_id}", current_user)
                elif verification_url:
                    # 如果没有截图，使用验证链接（兼容旧版本）
                    password_login_sessions[session_id]['status'] = 'verification_required'
                    password_login_sessions[session_id]['verification_url'] = verification_url
                    password_login_sessions[session_id]['screenshot_path'] = None
                    password_login_sessions[session_id]['qr_code_url'] = None
                    log_with_user('info', f"人脸认证验证链接已保存: {session_id}, URL: {verification_url}", current_user)
                    
                    # 发送通知到用户配置的渠道
                    def send_face_verification_notification():
                        """在后台线程中发送人脸验证通知"""
                        try:
                            from XianyuAutoAsync import XianyuLive
                            log_with_user('info', f"开始尝试发送人脸验证通知: {account_id}", current_user)
                            
                            # 尝试获取XianyuLive实例（如果账号已经存在）
                            live_instance = XianyuLive.get_instance(account_id)
                            
                            if live_instance:
                                log_with_user('info', f"找到账号实例，准备发送通知: {account_id}", current_user)
                                # 创建新的事件循环来运行异步通知
                                new_loop = asyncio.new_event_loop()
                                asyncio.set_event_loop(new_loop)
                                try:
                                    new_loop.run_until_complete(
                                        live_instance.send_token_refresh_notification(
                                            error_message=message,
                                            notification_type="face_verification",
                                            verification_url=verification_url
                                        )
                                    )
                                    log_with_user('info', f"✅ 已发送人脸验证通知: {account_id}", current_user)
                                except Exception as notify_err:
                                    log_with_user('error', f"发送人脸验证通知失败: {str(notify_err)}", current_user)
                                    import traceback
                                    log_with_user('error', f"通知错误详情: {traceback.format_exc()}", current_user)
                                finally:
                                    new_loop.close()
                            else:
                                # 如果账号实例不存在，记录警告并尝试从数据库获取通知配置
                                log_with_user('warning', f"账号实例不存在: {account_id}，尝试从数据库获取通知配置", current_user)
                                try:
                                    # 尝试从数据库获取通知配置
                                    notifications = db_manager.get_account_notifications(account_id)
                                    if notifications:
                                        log_with_user('info', f"找到 {len(notifications)} 个通知配置，但需要账号实例才能发送", current_user)
                                        log_with_user('warning', f"账号实例不存在，无法发送通知: {account_id}。请确保账号已登录并运行中。", current_user)
                                    else:
                                        log_with_user('warning', f"账号 {account_id} 未配置通知渠道", current_user)
                                except Exception as db_err:
                                    log_with_user('error', f"获取通知配置失败: {str(db_err)}", current_user)
                        except Exception as notify_err:
                            log_with_user('error', f"发送人脸验证通知时出错: {str(notify_err)}", current_user)
                            import traceback
                            log_with_user('error', f"通知错误详情: {traceback.format_exc()}", current_user)
                    
                    # 在后台线程中发送通知，避免阻塞登录流程
                    import threading
                    notification_thread = threading.Thread(target=send_face_verification_notification)
                    notification_thread.daemon = True
                    notification_thread.start()
                    log_with_user('info', f"已启动人脸验证通知发送线程: {account_id}", current_user)
            except Exception as e:
                log_with_user('error', f"处理人脸认证通知失败: {str(e)}", current_user)
        
        # 调用登录方法（同步方法，需要在后台线程中执行）
        import threading
        
        def run_login():
            try:
                cookies_dict = slider_instance.login_with_password_playwright(
                    account=account,
                    password=password,
                    show_browser=show_browser,
                    notification_callback=notification_callback
                )
                
                if cookies_dict is None:
                    password_login_sessions[session_id]['status'] = 'failed'
                    password_login_sessions[session_id]['error'] = '登录失败，请检查账号密码是否正确'
                    log_with_user('error', f"账号密码登录失败: {account_id}", current_user)
                    return
                
                # 将cookie字典转换为字符串格式
                cookies_str = '; '.join([f"{k}={v}" for k, v in cookies_dict.items()])
                
                log_with_user('info', f"账号密码登录成功，获取到 {len(cookies_dict)} 个Cookie字段: {account_id}", current_user)
                
                # 检查是否已存在相同账号ID的Cookie
                existing_cookies = db_manager.get_all_cookies(user_id)
                is_new_account = account_id not in existing_cookies
                
                # 保存账号密码和Cookie到数据库
                # 使用 update_cookie_account_info 来保存，它会自动处理新账号和现有账号的情况
                update_success = db_manager.update_cookie_account_info(
                    account_id,
                    cookie_value=cookies_str,
                    username=account,
                    password=password,
                    show_browser=show_browser,
                    user_id=user_id  # 新账号时需要提供user_id
                )
                
                if update_success:
                    if is_new_account:
                        log_with_user('info', f"新账号Cookie和账号密码已保存: {account_id}", current_user)
                    else:
                        log_with_user('info', f"现有账号Cookie和账号密码已更新: {account_id}", current_user)
                else:
                    log_with_user('error', f"保存账号信息失败: {account_id}", current_user)
                
                # 添加到或更新cookie_manager（注意：不要在这里调用add_cookie或update_cookie，因为它们会覆盖账号密码）
                # 账号密码已经在上面通过update_cookie_account_info保存了
                # 这里只需要更新内存中的cookie值，不保存到数据库（避免覆盖账号密码）
                if cookie_manager.manager:
                    # 更新内存中的cookie值
                    cookie_manager.manager.cookies[account_id] = cookies_str
                    log_with_user('info', f"已更新cookie_manager中的Cookie（内存）: {account_id}", current_user)
                    
                    # 如果是新账号，需要启动任务
                    if is_new_account:
                        # 使用异步方式启动任务，但不保存到数据库（避免覆盖账号密码）
                        try:
                            import asyncio
                            loop = cookie_manager.manager.loop
                            if loop:
                                # 确保关键词列表存在
                                if account_id not in cookie_manager.manager.keywords:
                                    cookie_manager.manager.keywords[account_id] = []
                                
                                # 在后台启动任务（使用线程安全的方式，因为run_login是在后台线程中运行的）
                                try:
                                    # 尝试使用run_coroutine_threadsafe，这是线程安全的方式
                                    fut = asyncio.run_coroutine_threadsafe(
                                        cookie_manager.manager._run_xianyu(account_id, cookies_str, user_id),
                                        loop
                                    )
                                    # 不等待结果，让它在后台运行
                                    log_with_user('info', f"已启动新账号任务: {account_id}", current_user)
                                except RuntimeError as e:
                                    # 如果事件循环未运行，记录警告但不影响登录成功
                                    log_with_user('warning', f"事件循环未运行，无法启动新账号任务: {account_id}, 错误: {str(e)}", current_user)
                                    log_with_user('info', f"账号已保存，将在系统重启后自动启动任务: {account_id}", current_user)
                        except Exception as task_err:
                            log_with_user('warning', f"启动新账号任务失败: {account_id}, 错误: {str(task_err)}", current_user)
                            import traceback
                            logger.error(traceback.format_exc())
                
                # 登录成功后，调用_refresh_cookies_via_browser刷新Cookie
                try:
                    log_with_user('info', f"开始调用_refresh_cookies_via_browser刷新Cookie: {account_id}", current_user)
                    from XianyuAutoAsync import XianyuLive
                    
                    # 创建临时的XianyuLive实例来刷新Cookie
                    temp_xianyu = XianyuLive(
                        cookies_str=cookies_str,
                        cookie_id=account_id,
                        user_id=user_id
                    )
                    
                    # 重置扫码登录Cookie刷新标志，确保账号密码登录后能立即刷新
                    try:
                        temp_xianyu.reset_qr_cookie_refresh_flag()
                        log_with_user('info', f"已重置扫码登录Cookie刷新标志: {account_id}", current_user)
                    except Exception as reset_err:
                        log_with_user('debug', f"重置扫码登录Cookie刷新标志失败（不影响刷新）: {str(reset_err)}", current_user)
                    
                    # 在后台异步执行刷新（不阻塞主流程）
                    async def refresh_cookies_task():
                        try:
                            refresh_success = await temp_xianyu._refresh_cookies_via_browser(triggered_by_refresh_token=False)
                            if refresh_success:
                                log_with_user('info', f"Cookie刷新成功: {account_id}", current_user)
                                # 刷新成功后，从数据库获取更新后的Cookie
                                updated_cookie_info = db_manager.get_cookie_details(account_id)
                                if updated_cookie_info:
                                    refreshed_cookies = updated_cookie_info.get('value', '')
                                    if refreshed_cookies:
                                        # 更新cookie_manager中的Cookie
                                        if cookie_manager.manager:
                                            cookie_manager.manager.update_cookie(account_id, refreshed_cookies, save_to_db=False)
                                        log_with_user('info', f"已更新刷新后的Cookie到cookie_manager: {account_id}", current_user)
                            else:
                                log_with_user('warning', f"Cookie刷新失败或跳过: {account_id}", current_user)
                        except Exception as refresh_e:
                            log_with_user('error', f"刷新Cookie时出错: {account_id}, 错误: {str(refresh_e)}", current_user)
                            import traceback
                            logger.error(traceback.format_exc())
                    
                    # 在后台线程中运行异步任务
                    # 由于run_login是在线程中运行的，需要创建新的事件循环
                    def run_async_refresh():
                        try:
                            import asyncio
                            # 创建新的事件循环
                            new_loop = asyncio.new_event_loop()
                            asyncio.set_event_loop(new_loop)
                            try:
                                new_loop.run_until_complete(refresh_cookies_task())
                            finally:
                                new_loop.close()
                        except Exception as e:
                            log_with_user('error', f"运行异步刷新任务失败: {account_id}, 错误: {str(e)}", current_user)
                    
                    # 在后台线程中执行刷新任务
                    refresh_thread = threading.Thread(target=run_async_refresh, daemon=True)
                    refresh_thread.start()
                    
                except Exception as refresh_err:
                    log_with_user('warning', f"调用_refresh_cookies_via_browser失败: {account_id}, 错误: {str(refresh_err)}", current_user)
                    # 刷新失败不影响登录成功
                
                # 更新会话状态
                password_login_sessions[session_id]['status'] = 'success'
                password_login_sessions[session_id]['account_id'] = account_id
                password_login_sessions[session_id]['is_new_account'] = is_new_account
                password_login_sessions[session_id]['cookie_count'] = len(cookies_dict)
                
            except Exception as e:
                error_msg = str(e)
                password_login_sessions[session_id]['status'] = 'failed'
                password_login_sessions[session_id]['error'] = error_msg
                log_with_user('error', f"账号密码登录失败: {account_id}, 错误: {error_msg}", current_user)
                logger.info(f"会话 {session_id} 状态已更新为 failed，错误消息: {error_msg}")  # 添加日志确认状态更新
                import traceback
                logger.error(traceback.format_exc())
            finally:
                # 清理实例（释放并发槽位）
                try:
                    from utils.xianyu_slider_stealth import concurrency_manager
                    concurrency_manager.unregister_instance(account_id)
                    log_with_user('debug', f"已释放并发槽位: {account_id}", current_user)
                except Exception as cleanup_e:
                    log_with_user('warning', f"清理实例时出错: {str(cleanup_e)}", current_user)
        
        # 在后台线程中执行登录
        login_thread = threading.Thread(target=run_login, daemon=True)
        login_thread.start()
        
    except Exception as e:
        password_login_sessions[session_id]['status'] = 'failed'
        password_login_sessions[session_id]['error'] = str(e)
        log_with_user('error', f"执行账号密码登录任务异常: {str(e)}", current_user)
        import traceback
        logger.error(traceback.format_exc())


@app.post("/password-login")
async def password_login(
    request: Dict[str, Any],
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """账号密码登录接口（异步，支持人脸认证）"""
    try:
        account_id = request.get('account_id')
        account = request.get('account')
        password = request.get('password')
        show_browser = request.get('show_browser', False)
        
        if not account_id or not account or not password:
            return {'success': False, 'message': '账号ID、登录账号和密码不能为空'}
        
        log_with_user('info', f"开始账号密码登录: {account_id}, 账号: {account}", current_user)
        
        # 生成会话ID
        import secrets
        session_id = secrets.token_urlsafe(16)
        
        user_id = current_user['user_id']
        
        # 创建登录会话
        password_login_sessions[session_id] = {
            'account_id': account_id,
            'account': account,
            'password': password,
            'show_browser': show_browser,
            'status': 'processing',
            'verification_url': None,
            'screenshot_path': None,
            'qr_code_url': None,
            'slider_instance': None,
            'task': None,
            'timestamp': time.time(),
            'user_id': user_id
        }
        
        # 启动后台登录任务
        task = asyncio.create_task(_execute_password_login(
            session_id, account_id, account, password, show_browser, user_id, current_user
        ))
        password_login_sessions[session_id]['task'] = task
        
        return {
            'success': True,
            'session_id': session_id,
            'status': 'processing',
            'message': '登录任务已启动，请等待...'
        }
        
    except Exception as e:
        log_with_user('error', f"账号密码登录异常: {str(e)}", current_user)
        import traceback
        logger.error(traceback.format_exc())
        return {'success': False, 'message': f'登录失败: {str(e)}'}


@app.get("/password-login/check/{session_id}")
async def check_password_login_status(
    session_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """检查账号密码登录状态"""
    try:
        # 清理过期会话（超过1小时）
        current_time = time.time()
        expired_sessions = [
            sid for sid, session in password_login_sessions.items()
            if current_time - session['timestamp'] > 3600
        ]
        for sid in expired_sessions:
            if sid in password_login_sessions:
                del password_login_sessions[sid]
        
        if session_id not in password_login_sessions:
            return {'status': 'not_found', 'message': '会话不存在或已过期'}
        
        session = password_login_sessions[session_id]
        
        # 检查用户权限
        if session['user_id'] != current_user['user_id']:
            return {'status': 'forbidden', 'message': '无权限访问该会话'}
        
        status = session['status']
        
        if status == 'verification_required':
            # 需要人脸认证
            screenshot_path = session.get('screenshot_path')
            verification_url = session.get('verification_url')
            return {
                'status': 'verification_required',
                'verification_url': verification_url,
                'screenshot_path': screenshot_path,
                'qr_code_url': session.get('qr_code_url'),  # 保留兼容性
                'message': '需要人脸验证，请查看验证截图' if screenshot_path else '需要人脸验证，请点击验证链接'
            }
        elif status == 'success':
            # 登录成功
            # 删除截图（如果存在）
            screenshot_path = session.get('screenshot_path')
            if screenshot_path:
                try:
                    from utils.image_utils import image_manager
                    if image_manager.delete_image(screenshot_path):
                        log_with_user('info', f"验证成功后已删除截图: {screenshot_path}", current_user)
                    else:
                        log_with_user('warning', f"删除截图失败: {screenshot_path}", current_user)
                except Exception as e:
                    log_with_user('error', f"删除截图时出错: {str(e)}", current_user)
            
            result = {
                'status': 'success',
                'message': f'账号 {session["account_id"]} 登录成功',
                'account_id': session['account_id'],
                'is_new_account': session.get('is_new_account', False),
                'cookie_count': session.get('cookie_count', 0)
            }
            # 清理会话
            del password_login_sessions[session_id]
            return result
        elif status == 'failed':
            # 登录失败
            # 删除截图（如果存在）
            screenshot_path = session.get('screenshot_path')
            if screenshot_path:
                try:
                    from utils.image_utils import image_manager
                    if image_manager.delete_image(screenshot_path):
                        log_with_user('info', f"验证失败后已删除截图: {screenshot_path}", current_user)
                    else:
                        log_with_user('warning', f"删除截图失败: {screenshot_path}", current_user)
                except Exception as e:
                    log_with_user('error', f"删除截图时出错: {str(e)}", current_user)
            
            error_msg = session.get('error', '登录失败')
            log_with_user('info', f"返回登录失败状态: {session_id}, 错误消息: {error_msg}", current_user)  # 添加日志
            result = {
                'status': 'failed',
                'message': error_msg,
                'error': error_msg  # 也包含error字段，确保前端能获取到
            }
            # 清理会话
            del password_login_sessions[session_id]
            return result
        else:
            # 处理中
            return {
                'status': 'processing',
                'message': '登录处理中，请稍候...'
            }
        
    except Exception as e:
        log_with_user('error', f"检查账号密码登录状态异常: {str(e)}", current_user)
        return {'status': 'error', 'message': str(e)}


# ========================= 人脸验证截图相关接口 =========================

@app.get("/face-verification/screenshot/{account_id}")
async def get_account_face_verification_screenshot(
    account_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """获取指定账号的人脸验证截图"""
    try:
        import glob
        from datetime import datetime
        
        # 检查账号是否属于当前用户
        user_id = current_user['user_id']
        username = current_user['username']
        
        # 如果是管理员，允许访问所有账号
        is_admin = bool(current_user.get('is_admin', False))
        
        if not is_admin:
            cookie_info = db_manager.get_cookie_details(account_id)
            if not cookie_info:
                log_with_user('warning', f"账号 {account_id} 不存在", current_user)
                return {
                    'success': False,
                    'message': '账号不存在'
                }
            
            cookie_user_id = cookie_info.get('user_id')
            if cookie_user_id != user_id:
                log_with_user('warning', f"用户 {user_id} 尝试访问账号 {account_id}（归属用户: {cookie_user_id}）", current_user)
                return {
                    'success': False,
                    'message': '无权访问该账号'
                }
        
        # 获取该账号的验证截图
        screenshots_dir = os.path.join(static_dir, 'uploads', 'images')
        pattern = os.path.join(screenshots_dir, f'face_verify_{account_id}_*.jpg')
        screenshot_files = glob.glob(pattern)
        
        log_with_user('debug', f"查找截图: {pattern}, 找到 {len(screenshot_files)} 个文件", current_user)
        
        if not screenshot_files:
            log_with_user('warning', f"账号 {account_id} 没有找到验证截图", current_user)
            return {
                'success': False,
                'message': '未找到验证截图'
            }
        
        # 获取最新的截图
        latest_file = max(screenshot_files, key=os.path.getmtime)
        filename = os.path.basename(latest_file)
        stat = os.stat(latest_file)
        
        screenshot_info = {
            'filename': filename,
            'account_id': account_id,
            'path': f'/static/uploads/images/{filename}',
            'size': stat.st_size,
            'created_time': stat.st_ctime,
            'created_time_str': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
        }
        
        log_with_user('info', f"获取账号 {account_id} 的验证截图", current_user)
        
        return {
            'success': True,
            'screenshot': screenshot_info
        }
        
    except Exception as e:
        log_with_user('error', f"获取验证截图失败: {str(e)}", current_user)
        return {
            'success': False,
            'message': str(e)
        }


@app.delete("/face-verification/screenshot/{account_id}")
async def delete_account_face_verification_screenshot(
    account_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """删除指定账号的人脸验证截图"""
    try:
        import glob
        
        # 检查账号是否属于当前用户
        user_id = current_user['user_id']
        cookie_info = db_manager.get_cookie_details(account_id)
        if not cookie_info or cookie_info.get('user_id') != user_id:
            return {
                'success': False,
                'message': '无权访问该账号'
            }
        
        # 删除该账号的所有验证截图
        screenshots_dir = os.path.join(static_dir, 'uploads', 'images')
        pattern = os.path.join(screenshots_dir, f'face_verify_{account_id}_*.jpg')
        screenshot_files = glob.glob(pattern)
        
        deleted_count = 0
        for file_path in screenshot_files:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    deleted_count += 1
                    log_with_user('info', f"删除账号 {account_id} 的验证截图: {os.path.basename(file_path)}", current_user)
            except Exception as e:
                log_with_user('error', f"删除截图失败 {file_path}: {str(e)}", current_user)
        
        return {
            'success': True,
            'message': f'已删除 {deleted_count} 个验证截图',
            'deleted_count': deleted_count
        }
        
    except Exception as e:
        log_with_user('error', f"删除验证截图失败: {str(e)}", current_user)
        return {
            'success': False,
            'message': str(e)
        }


# ========================= 扫码登录相关接口 =========================

@app.post("/qr-login/generate")
async def generate_qr_code(current_user: Dict[str, Any] = Depends(get_current_user)):
    """生成扫码登录二维码"""
    try:
        log_with_user('info', "请求生成扫码登录二维码", current_user)

        result = await qr_login_manager.generate_qr_code()

        if result['success']:
            log_with_user('info', f"扫码登录二维码生成成功: {result['session_id']}", current_user)
        else:
            log_with_user('warning', f"扫码登录二维码生成失败: {result.get('message', '未知错误')}", current_user)

        return result

    except Exception as e:
        log_with_user('error', f"生成扫码登录二维码异常: {str(e)}", current_user)
        return {'success': False, 'message': f'生成二维码失败: {str(e)}'}


@app.get("/qr-login/check/{session_id}")
async def check_qr_code_status(session_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """检查扫码登录状态"""
    try:
        # 清理过期记录
        cleanup_qr_check_records()

        # 检查是否已经处理过
        if session_id in qr_check_processed:
            record = qr_check_processed[session_id]
            if record['processed']:
                log_with_user('debug', f"扫码登录session {session_id} 已处理过，直接返回", current_user)
                # 返回简单的成功状态，避免重复处理
                return {'status': 'already_processed', 'message': '该会话已处理完成'}

        # 获取该session的锁
        session_lock = qr_check_locks[session_id]

        # 使用非阻塞方式尝试获取锁
        if session_lock.locked():
            log_with_user('debug', f"扫码登录session {session_id} 正在被其他请求处理，跳过", current_user)
            return {'status': 'processing', 'message': '正在处理中，请稍候...'}

        async with session_lock:
            # 再次检查是否已处理（双重检查）
            if session_id in qr_check_processed and qr_check_processed[session_id]['processed']:
                log_with_user('debug', f"扫码登录session {session_id} 在获取锁后发现已处理，直接返回", current_user)
                return {'status': 'already_processed', 'message': '该会话已处理完成'}

            # 清理过期会话
            qr_login_manager.cleanup_expired_sessions()

            # 获取会话状态
            status_info = qr_login_manager.get_session_status(session_id)
            log_with_user('info', f"获取会话状态1111111: {status_info}", current_user)
            if status_info['status'] == 'success':
                log_with_user('info', f"获取会话状态22222222: {status_info}", current_user)
                # 登录成功，处理Cookie（现在包含获取真实cookie的逻辑）
                cookies_info = qr_login_manager.get_session_cookies(session_id)
                log_with_user('info', f"获取会话Cookie: {cookies_info}", current_user)
                if cookies_info:
                    account_info = await process_qr_login_cookies(
                        cookies_info['cookies'],
                        cookies_info['unb'],
                        current_user
                    )
                    status_info['account_info'] = account_info

                    log_with_user('info', f"扫码登录处理完成: {session_id}, 账号: {account_info.get('account_id', 'unknown')}", current_user)

                    # 标记该session已处理
                    qr_check_processed[session_id] = {
                        'processed': True,
                        'timestamp': time.time()
                    }

            return status_info

    except Exception as e:
        log_with_user('error', f"检查扫码登录状态异常: {str(e)}", current_user)
        return {'status': 'error', 'message': str(e)}


async def process_qr_login_cookies(cookies: str, unb: str, current_user: Dict[str, Any]) -> Dict[str, Any]:
    """处理扫码登录获取的Cookie - 先获取真实cookie再保存到数据库"""
    try:
        user_id = current_user['user_id']

        # 检查是否已存在相同unb的账号
        existing_cookies = db_manager.get_all_cookies(user_id)
        existing_account_id = None

        for account_id, cookie_value in existing_cookies.items():
            try:
                # 解析现有Cookie中的unb
                existing_cookie_dict = trans_cookies(cookie_value)
                if existing_cookie_dict.get('unb') == unb:
                    existing_account_id = account_id
                    break
            except:
                continue

        # 确定账号ID
        if existing_account_id:
            account_id = existing_account_id
            is_new_account = False
            log_with_user('info', f"扫码登录找到现有账号: {account_id}, UNB: {unb}", current_user)
        else:
            # 创建新账号，使用unb作为账号ID
            account_id = unb

            # 确保账号ID唯一
            counter = 1
            original_account_id = account_id
            while account_id in existing_cookies:
                account_id = f"{original_account_id}_{counter}"
                counter += 1

            is_new_account = True
            log_with_user('info', f"扫码登录准备创建新账号: {account_id}, UNB: {unb}", current_user)

        # 第一步：使用扫码cookie获取真实cookie
        log_with_user('info', f"开始使用扫码cookie获取真实cookie: {account_id}", current_user)

        try:
            # 创建一个临时的XianyuLive实例来执行cookie刷新
            from XianyuAutoAsync import XianyuLive

            # 使用扫码登录的cookie创建临时实例
            temp_instance = XianyuLive(
                cookies_str=cookies,
                cookie_id=account_id,
                user_id=user_id
            )

            # 执行cookie刷新获取真实cookie
            refresh_success = await temp_instance.refresh_cookies_from_qr_login(
                qr_cookies_str=cookies,
                cookie_id=account_id,
                user_id=user_id
            )

            if refresh_success:
                log_with_user('info', f"扫码登录真实cookie获取成功: {account_id}", current_user)

                # 从数据库获取刚刚保存的真实cookie
                updated_cookie_info = db_manager.get_cookie_by_id(account_id)
                if updated_cookie_info:
                    real_cookies = updated_cookie_info['cookies_str']
                    log_with_user('info', f"已获取真实cookie，长度: {len(real_cookies)}", current_user)

                    # 第二步：将真实cookie添加到cookie_manager（如果是新账号）或更新现有账号
                    if cookie_manager.manager:
                        if is_new_account:
                            cookie_manager.manager.add_cookie(account_id, real_cookies)
                            log_with_user('info', f"已将真实cookie添加到cookie_manager: {account_id}", current_user)
                        else:
                            # refresh_cookies_from_qr_login 已经保存到数据库了，这里不需要再保存
                            cookie_manager.manager.update_cookie(account_id, real_cookies, save_to_db=False)
                            log_with_user('info', f"已更新cookie_manager中的真实cookie: {account_id}", current_user)

                    return {
                        'account_id': account_id,
                        'is_new_account': is_new_account,
                        'real_cookie_refreshed': True,
                        'cookie_length': len(real_cookies)
                    }
                else:
                    log_with_user('error', f"无法从数据库获取真实cookie: {account_id}", current_user)
                    # 降级处理：使用原始扫码cookie
                    return await _fallback_save_qr_cookie(account_id, cookies, user_id, is_new_account, current_user, "无法从数据库获取真实cookie")
            else:
                log_with_user('warning', f"扫码登录真实cookie获取失败: {account_id}", current_user)
                # 降级处理：使用原始扫码cookie
                return await _fallback_save_qr_cookie(account_id, cookies, user_id, is_new_account, current_user, "真实cookie获取失败")

        except Exception as refresh_e:
            log_with_user('error', f"扫码登录真实cookie获取异常: {str(refresh_e)}", current_user)
            # 降级处理：使用原始扫码cookie
            return await _fallback_save_qr_cookie(account_id, cookies, user_id, is_new_account, current_user, f"获取真实cookie异常: {str(refresh_e)}")

    except Exception as e:
        log_with_user('error', f"处理扫码登录Cookie失败: {str(e)}", current_user)
        raise e


async def _fallback_save_qr_cookie(account_id: str, cookies: str, user_id: int, is_new_account: bool, current_user: Dict[str, Any], error_reason: str) -> Dict[str, Any]:
    """降级处理：当无法获取真实cookie时，保存原始扫码cookie"""
    try:
        log_with_user('warning', f"降级处理 - 保存原始扫码cookie: {account_id}, 原因: {error_reason}", current_user)

        # 保存原始扫码cookie到数据库
        if is_new_account:
            db_manager.save_cookie(account_id, cookies, user_id)
            log_with_user('info', f"降级处理 - 新账号原始cookie已保存: {account_id}", current_user)
        else:
            # 现有账号使用 update_cookie_account_info 避免覆盖其他字段
            db_manager.update_cookie_account_info(account_id, cookie_value=cookies)
            log_with_user('info', f"降级处理 - 现有账号原始cookie已更新: {account_id}", current_user)

        # 添加到或更新cookie_manager
        if cookie_manager.manager:
            if is_new_account:
                cookie_manager.manager.add_cookie(account_id, cookies)
                log_with_user('info', f"降级处理 - 已将原始cookie添加到cookie_manager: {account_id}", current_user)
            else:
                # update_cookie_account_info 已经保存到数据库了，这里不需要再保存
                cookie_manager.manager.update_cookie(account_id, cookies, save_to_db=False)
                log_with_user('info', f"降级处理 - 已更新cookie_manager中的原始cookie: {account_id}", current_user)

        return {
            'account_id': account_id,
            'is_new_account': is_new_account,
            'real_cookie_refreshed': False,
            'fallback_reason': error_reason,
            'cookie_length': len(cookies)
        }

    except Exception as fallback_e:
        log_with_user('error', f"降级处理失败: {str(fallback_e)}", current_user)
        raise fallback_e


@app.post("/qr-login/refresh-cookies")
async def refresh_cookies_from_qr_login(
    request: Dict[str, Any],
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """使用扫码登录获取的cookie访问指定界面获取真实cookie并存入数据库"""
    try:
        qr_cookies = request.get('qr_cookies')
        cookie_id = request.get('cookie_id')

        if not qr_cookies:
            return {'success': False, 'message': '缺少扫码登录cookie'}

        if not cookie_id:
            return {'success': False, 'message': '缺少cookie_id'}

        log_with_user('info', f"开始使用扫码cookie刷新真实cookie: {cookie_id}", current_user)

        # 创建一个临时的XianyuLive实例来执行cookie刷新
        from XianyuAutoAsync import XianyuLive

        # 使用扫码登录的cookie创建临时实例
        temp_instance = XianyuLive(
            cookies_str=qr_cookies,
            cookie_id=cookie_id,
            user_id=current_user['user_id']
        )

        # 执行cookie刷新
        success = await temp_instance.refresh_cookies_from_qr_login(
            qr_cookies_str=qr_cookies,
            cookie_id=cookie_id,
            user_id=current_user['user_id']
        )

        if success:
            log_with_user('info', f"扫码cookie刷新成功: {cookie_id}", current_user)

            # 如果cookie_manager存在，更新其中的cookie
            if cookie_manager.manager:
                # 从数据库获取更新后的cookie
                updated_cookie_info = db_manager.get_cookie_by_id(cookie_id)
                if updated_cookie_info:
                    # refresh_cookies_from_qr_login 已经保存到数据库了，这里不需要再保存
                    cookie_manager.manager.update_cookie(cookie_id, updated_cookie_info['cookies_str'], save_to_db=False)
                    log_with_user('info', f"已更新cookie_manager中的cookie: {cookie_id}", current_user)

            return {
                'success': True,
                'message': '真实cookie获取并保存成功',
                'cookie_id': cookie_id
            }
        else:
            log_with_user('error', f"扫码cookie刷新失败: {cookie_id}", current_user)
            return {'success': False, 'message': '获取真实cookie失败'}

    except Exception as e:
        log_with_user('error', f"扫码cookie刷新异常: {str(e)}", current_user)
        return {'success': False, 'message': f'刷新cookie失败: {str(e)}'}


@app.post("/qr-login/reset-cooldown/{cookie_id}")
async def reset_qr_cookie_refresh_cooldown(
    cookie_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """重置指定账号的扫码登录Cookie刷新冷却时间"""
    try:
        log_with_user('info', f"重置扫码登录Cookie刷新冷却时间: {cookie_id}", current_user)

        # 检查cookie是否存在
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {'success': False, 'message': '账号不存在'}

        # 如果cookie_manager中有对应的实例，直接重置
        if cookie_manager.manager and cookie_id in cookie_manager.manager.instances:
            instance = cookie_manager.manager.instances[cookie_id]
            remaining_time_before = instance.get_qr_cookie_refresh_remaining_time()
            instance.reset_qr_cookie_refresh_flag()

            log_with_user('info', f"已重置账号 {cookie_id} 的扫码登录冷却时间，原剩余时间: {remaining_time_before}秒", current_user)

            return {
                'success': True,
                'message': '扫码登录Cookie刷新冷却时间已重置',
                'cookie_id': cookie_id,
                'previous_remaining_time': remaining_time_before
            }
        else:
            # 如果没有活跃实例，返回成功（因为没有冷却时间需要重置）
            log_with_user('info', f"账号 {cookie_id} 没有活跃实例，无需重置冷却时间", current_user)
            return {
                'success': True,
                'message': '账号没有活跃实例，无需重置冷却时间',
                'cookie_id': cookie_id
            }

    except Exception as e:
        log_with_user('error', f"重置扫码登录冷却时间异常: {str(e)}", current_user)
        return {'success': False, 'message': f'重置冷却时间失败: {str(e)}'}


@app.get("/qr-login/cooldown-status/{cookie_id}")
async def get_qr_cookie_refresh_cooldown_status(
    cookie_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """获取指定账号的扫码登录Cookie刷新冷却状态"""
    try:
        # 检查cookie是否存在
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {'success': False, 'message': '账号不存在'}

        # 如果cookie_manager中有对应的实例，获取冷却状态
        if cookie_manager.manager and cookie_id in cookie_manager.manager.instances:
            instance = cookie_manager.manager.instances[cookie_id]
            remaining_time = instance.get_qr_cookie_refresh_remaining_time()
            cooldown_duration = instance.qr_cookie_refresh_cooldown
            last_refresh_time = instance.last_qr_cookie_refresh_time

            return {
                'success': True,
                'cookie_id': cookie_id,
                'remaining_time': remaining_time,
                'cooldown_duration': cooldown_duration,
                'last_refresh_time': last_refresh_time,
                'is_in_cooldown': remaining_time > 0,
                'remaining_minutes': remaining_time // 60,
                'remaining_seconds': remaining_time % 60
            }
        else:
            return {
                'success': True,
                'cookie_id': cookie_id,
                'remaining_time': 0,
                'cooldown_duration': 600,  # 默认10分钟
                'last_refresh_time': 0,
                'is_in_cooldown': False,
                'message': '账号没有活跃实例'
            }

    except Exception as e:
        log_with_user('error', f"获取扫码登录冷却状态异常: {str(e)}", current_user)
        return {'success': False, 'message': f'获取冷却状态失败: {str(e)}'}


@app.put('/cookies/{cid}/status')
def update_cookie_status(cid: str, status_data: CookieStatusIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号的启用/禁用状态"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        cookie_manager.manager.update_cookie_status(cid, status_data.enabled)
        return {'msg': 'status updated', 'enabled': status_data.enabled}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ------------------------- 默认回复管理接口 -------------------------

@app.get('/default-replies/{cid}')
def get_default_reply(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        result = db_manager.get_default_reply(cid)
        if result is None:
            # 如果没有设置，返回默认值
            return {'enabled': False, 'reply_content': '', 'reply_once': False}
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put('/default-replies/{cid}')
def update_default_reply(cid: str, reply_data: DefaultReplyIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        db_manager.save_default_reply(cid, reply_data.enabled, reply_data.reply_content, reply_data.reply_once, reply_data.reply_image_url)
        return {'msg': 'default reply updated', 'enabled': reply_data.enabled, 'reply_once': reply_data.reply_once, 'reply_image_url': reply_data.reply_image_url}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/default-replies')
def get_all_default_replies(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 只返回当前用户的默认回复设置
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        all_replies = db_manager.get_all_default_replies()
        # 过滤只属于当前用户的回复设置
        user_replies = {cid: reply for cid, reply in all_replies.items() if cid in user_cookies}
        return user_replies
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/default-replies/{cid}')
def delete_default_reply(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.delete_default_reply(cid)
        if success:
            return {'msg': 'default reply deleted'}
        else:
            raise HTTPException(status_code=400, detail='删除失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/default-replies/{cid}/clear-records')
def clear_default_reply_records(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """清空指定账号的默认回复记录"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        db_manager.clear_default_reply_records(cid)
        return {'msg': 'default reply records cleared'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 默认回复管理接口（单数形式兼容路由） -------------------------
# 兼容前端使用 /api/default-reply/ 的请求

@app.get('/api/default-reply/{cid}')
def get_default_reply_compat(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的默认回复设置（兼容路由）"""
    return get_default_reply(cid, current_user)


@app.put('/api/default-reply/{cid}')
def update_default_reply_compat(cid: str, reply_data: DefaultReplyIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的默认回复设置（兼容路由）"""
    return update_default_reply(cid, reply_data, current_user)


@app.delete('/api/default-reply/{cid}')
def delete_default_reply_compat(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除指定账号的默认回复设置（兼容路由）"""
    return delete_default_reply(cid, current_user)


@app.post('/api/default-reply/{cid}/clear-records')
def clear_default_reply_records_compat(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """清空指定账号的默认回复记录（兼容路由）"""
    return clear_default_reply_records(cid, current_user)


# ------------------------- 通知渠道管理接口 -------------------------

@app.get('/notification-channels')
def get_notification_channels(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取所有通知渠道"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        return db_manager.get_notification_channels(user_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/notification-channels')
def create_notification_channel(channel_data: NotificationChannelIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建通知渠道"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        channel_id = db_manager.create_notification_channel(
            channel_data.name,
            channel_data.type,
            channel_data.config,
            user_id
        )
        return {'msg': 'notification channel created', 'id': channel_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get('/notification-channels/{channel_id}')
def get_notification_channel(channel_id: int, _: None = Depends(require_auth)):
    """获取指定通知渠道"""
    from db_manager import db_manager
    try:
        channel = db_manager.get_notification_channel(channel_id)
        if not channel:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
        return channel
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put('/notification-channels/{channel_id}')
def update_notification_channel(channel_id: int, channel_data: NotificationChannelUpdate, _: None = Depends(require_auth)):
    """更新通知渠道"""
    from db_manager import db_manager
    try:
        success = db_manager.update_notification_channel(
            channel_id,
            channel_data.name,
            channel_data.config,
            channel_data.enabled
        )
        if success:
            return {'msg': 'notification channel updated'}
        else:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.delete('/notification-channels/{channel_id}')
def delete_notification_channel(channel_id: int, _: None = Depends(require_auth)):
    """删除通知渠道"""
    from db_manager import db_manager
    try:
        success = db_manager.delete_notification_channel(channel_id)
        if success:
            return {'msg': 'notification channel deleted'}
        else:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 消息通知配置接口 -------------------------

@app.get('/message-notifications')
def get_all_message_notifications(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的消息通知配置"""
    from db_manager import db_manager
    try:
        # 只返回当前用户的消息通知配置
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        all_notifications = db_manager.get_all_message_notifications()
        # 过滤只属于当前用户的通知配置
        user_notifications = {cid: notifications for cid, notifications in all_notifications.items() if cid in user_cookies}
        return user_notifications
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/message-notifications/{cid}')
def get_account_notifications(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的消息通知配置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        return db_manager.get_account_notifications(cid)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/message-notifications/{cid}')
def set_message_notification(cid: str, notification_data: MessageNotificationIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """设置账号的消息通知"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 检查通知渠道是否存在
        channel = db_manager.get_notification_channel(notification_data.channel_id)
        if not channel:
            raise HTTPException(status_code=404, detail='通知渠道不存在')

        success = db_manager.set_message_notification(cid, notification_data.channel_id, notification_data.enabled)
        if success:
            return {'msg': 'message notification set'}
        else:
            raise HTTPException(status_code=400, detail='设置失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/message-notifications/account/{cid}')
def delete_account_notifications(cid: str, _: None = Depends(require_auth)):
    """删除账号的所有消息通知配置"""
    from db_manager import db_manager
    try:
        success = db_manager.delete_account_notifications(cid)
        if success:
            return {'msg': 'account notifications deleted'}
        else:
            raise HTTPException(status_code=404, detail='账号通知配置不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/message-notifications/{notification_id}')
def delete_message_notification(notification_id: int, _: None = Depends(require_auth)):
    """删除消息通知配置"""
    from db_manager import db_manager
    try:
        success = db_manager.delete_message_notification(notification_id)
        if success:
            return {'msg': 'message notification deleted'}
        else:
            raise HTTPException(status_code=404, detail='通知配置不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 系统设置接口 -------------------------

@app.get('/system-settings/public')
def get_public_system_settings():
    """获取公开的系统设置（无需认证）"""
    from db_manager import db_manager
    try:
        all_settings = db_manager.get_all_system_settings()
        # 只返回公开的配置项
        public_keys = {"registration_enabled", "show_default_login_info", "login_captcha_enabled"}
        return {k: v for k, v in all_settings.items() if k in public_keys}
    except Exception as e:
        logger.error(f"获取公开系统设置失败: {e}")
        # 返回默认值
        return {
            "registration_enabled": "true",
            "show_default_login_info": "true",
            "login_captcha_enabled": "true"
        }


@app.get('/system-settings')
def get_system_settings(_: None = Depends(require_auth)):
    """获取系统设置（排除敏感信息）"""
    from db_manager import db_manager
    try:
        settings = db_manager.get_all_system_settings()
        # 移除敏感信息
        if 'admin_password_hash' in settings:
            del settings['admin_password_hash']
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put('/system-settings/{key}')
def update_system_setting(key: str, setting_data: SystemSettingIn, _: None = Depends(require_auth)):
    """更新系统设置"""
    from db_manager import db_manager
    try:
        # 禁止直接修改密码哈希
        if key == 'admin_password_hash':
            raise HTTPException(status_code=400, detail='请使用密码修改接口')

        success = db_manager.set_system_setting(key, setting_data.value, setting_data.description)
        if success:
            return {'msg': 'system setting updated'}
        else:
            raise HTTPException(status_code=400, detail='更新失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 注册设置接口 -------------------------

@app.get('/registration-status')
def get_registration_status():
    """获取注册开关状态（公开接口，无需认证）"""
    from db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('registration_enabled')
        logger.info(f"从数据库获取的注册设置值: '{enabled_str}'")  # 调试信息

        # 如果设置不存在，默认为开启
        if enabled_str is None:
            enabled_bool = True
            message = '注册功能已开启'
        else:
            enabled_bool = enabled_str == 'true'
            message = '注册功能已开启' if enabled_bool else '注册功能已关闭'

        logger.info(f"解析后的注册状态: enabled={enabled_bool}, message='{message}'")  # 调试信息

        return {
            'enabled': enabled_bool,
            'message': message
        }
    except Exception as e:
        logger.error(f"获取注册状态失败: {e}")
        return {'enabled': True, 'message': '注册功能已开启'}  # 出错时默认开启


@app.get('/login-info-status')
def get_login_info_status():
    """获取默认登录信息显示状态（公开接口，无需认证）"""
    from db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('show_default_login_info')
        logger.debug(f"从数据库获取的登录信息显示设置值: '{enabled_str}'")

        # 如果设置不存在，默认为开启
        if enabled_str is None:
            enabled_bool = True
        else:
            enabled_bool = enabled_str == 'true'

        return {"enabled": enabled_bool}
    except Exception as e:
        logger.error(f"获取登录信息显示状态失败: {e}")
        # 出错时默认为开启
        return {"enabled": True}


class RegistrationSettingUpdate(BaseModel):
    enabled: bool


class LoginInfoSettingUpdate(BaseModel):
    enabled: bool


@app.put('/registration-settings')
def update_registration_settings(setting_data: RegistrationSettingUpdate, admin_user: Dict[str, Any] = Depends(require_admin)):
    """更新注册开关设置（仅管理员）"""
    from db_manager import db_manager
    try:
        enabled = setting_data.enabled
        success = db_manager.set_system_setting(
            'registration_enabled',
            'true' if enabled else 'false',
            '是否开启用户注册'
        )
        if success:
            log_with_user('info', f"更新注册设置: {'开启' if enabled else '关闭'}", admin_user)
            return {
                'success': True,
                'enabled': enabled,
                'message': f"注册功能已{'开启' if enabled else '关闭'}"
            }
        else:
            raise HTTPException(status_code=500, detail='更新注册设置失败')
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新注册设置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put('/login-info-settings')
def update_login_info_settings(setting_data: LoginInfoSettingUpdate, admin_user: Dict[str, Any] = Depends(require_admin)):
    """更新默认登录信息显示设置（仅管理员）"""
    from db_manager import db_manager
    try:
        enabled = setting_data.enabled
        success = db_manager.set_system_setting(
            'show_default_login_info',
            'true' if enabled else 'false',
            '是否显示默认登录信息'
        )
        if success:
            log_with_user('info', f"更新登录信息显示设置: {'开启' if enabled else '关闭'}", admin_user)
            return {
                'success': True,
                'enabled': enabled,
                'message': f"默认登录信息显示已{'开启' if enabled else '关闭'}"
            }
        else:
            raise HTTPException(status_code=500, detail='更新登录信息显示设置失败')
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新登录信息显示设置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))




@app.delete("/cookies/{cid}")
def remove_cookie(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        cookie_manager.manager.remove_cookie(cid)
        return {"msg": "removed"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class AutoConfirmUpdate(BaseModel):
    auto_confirm: bool


class RemarkUpdate(BaseModel):
    remark: str


class PauseDurationUpdate(BaseModel):
    pause_duration: int


@app.put("/cookies/{cid}/auto-confirm")
def update_auto_confirm(cid: str, update_data: AutoConfirmUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号的自动确认发货设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新数据库中的auto_confirm设置
        success = db_manager.update_auto_confirm(cid, update_data.auto_confirm)
        if not success:
            raise HTTPException(status_code=500, detail="更新自动确认发货设置失败")

        # 通知CookieManager更新设置（如果账号正在运行）
        if hasattr(cookie_manager.manager, 'update_auto_confirm_setting'):
            cookie_manager.manager.update_auto_confirm_setting(cid, update_data.auto_confirm)

        return {
            "msg": "success",
            "auto_confirm": update_data.auto_confirm,
            "message": f"自动确认发货已{'开启' if update_data.auto_confirm else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/auto-confirm")
def get_auto_confirm(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号的自动确认发货设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取auto_confirm设置
        auto_confirm = db_manager.get_auto_confirm(cid)
        return {
            "auto_confirm": auto_confirm,
            "message": f"自动确认发货当前{'开启' if auto_confirm else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cookies/{cid}/remark")
def update_cookie_remark(cid: str, update_data: RemarkUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号备注"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新备注
        success = db_manager.update_cookie_remark(cid, update_data.remark)
        if success:
            log_with_user('info', f"更新账号备注: {cid} -> {update_data.remark}", current_user)
            return {
                "message": "备注更新成功",
                "remark": update_data.remark
            }
        else:
            raise HTTPException(status_code=500, detail="备注更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/remark")
def get_cookie_remark(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号备注"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取Cookie详细信息（包含备注）
        cookie_details = db_manager.get_cookie_details(cid)
        if cookie_details:
            return {
                "remark": cookie_details.get('remark', ''),
                "message": "获取备注成功"
            }
        else:
            raise HTTPException(status_code=404, detail="账号不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cookies/{cid}/pause-duration")
def update_cookie_pause_duration(cid: str, update_data: PauseDurationUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号自动回复暂停时间"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 验证暂停时间范围（0-120分钟，0表示不暂停）
        if not (0 <= update_data.pause_duration <= 120):
            raise HTTPException(status_code=400, detail="暂停时间必须在0-120分钟之间（0表示不暂停）")

        # 更新暂停时间
        success = db_manager.update_cookie_pause_duration(cid, update_data.pause_duration)
        if success:
            log_with_user('info', f"更新账号自动回复暂停时间: {cid} -> {update_data.pause_duration}分钟", current_user)
            return {
                "message": "暂停时间更新成功",
                "pause_duration": update_data.pause_duration
            }
        else:
            raise HTTPException(status_code=500, detail="暂停时间更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/pause-duration")
def get_cookie_pause_duration(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号自动回复暂停时间"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取暂停时间
        pause_duration = db_manager.get_cookie_pause_duration(cid)
        return {
            "pause_duration": pause_duration,
            "message": "获取暂停时间成功"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class KeywordIn(BaseModel):
    keywords: Dict[str, str]  # key -> reply

class KeywordWithItemIdIn(BaseModel):
    keywords: List[Dict[str, Any]]  # [{"keyword": str, "reply": str, "item_id": str}]


@app.get("/keywords/{cid}")
def get_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 直接从数据库获取所有关键词（避免重复计算）
    item_keywords = db_manager.get_keywords_with_item_id(cid)

    # 转换为统一格式
    all_keywords = []
    for keyword, reply, item_id in item_keywords:
        all_keywords.append({
            "keyword": keyword,
            "reply": reply,
            "item_id": item_id,
            "type": "item" if item_id else "normal"
        })

    return all_keywords


@app.get("/keywords-with-item-id/{cid}")
def get_keywords_with_item_id(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 获取包含类型信息的关键词
    keywords = db_manager.get_keywords_with_type(cid)

    # 转换为前端需要的格式
    result = []
    for keyword_data in keywords:
        result.append({
            "keyword": keyword_data['keyword'],
            "reply": keyword_data['reply'],
            "item_id": keyword_data['item_id'] or "",
            "type": keyword_data['type'],
            "image_url": keyword_data['image_url']
        })

    return result


@app.post("/keywords/{cid}")
def update_keywords(cid: str, body: KeywordIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    kw_list = [(k, v) for k, v in body.keywords.items()]
    log_with_user('info', f"更新Cookie关键字: {cid}, 数量: {len(kw_list)}", current_user)

    cookie_manager.manager.update_keywords(cid, kw_list)
    log_with_user('info', f"Cookie关键字更新成功: {cid}", current_user)
    return {"msg": "updated", "count": len(kw_list)}


@app.post("/keywords-with-item-id/{cid}")
def update_keywords_with_item_id(cid: str, body: KeywordWithItemIdIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    # 验证数据格式
    keywords_to_save = []
    keyword_set = set()  # 用于检查当前提交的关键词中是否有重复

    for kw_data in body.keywords:
        keyword = kw_data.get('keyword', '').strip()
        reply = kw_data.get('reply', '').strip()
        item_id = kw_data.get('item_id', '').strip() or None

        if not keyword:
            raise HTTPException(status_code=400, detail="关键词不能为空")

        # 检查当前提交的关键词中是否有重复
        keyword_key = f"{keyword}|{item_id or ''}"
        if keyword_key in keyword_set:
            item_id_text = f"（商品ID: {item_id}）" if item_id else "（通用关键词）"
            raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' {item_id_text} 在当前提交中重复")
        keyword_set.add(keyword_key)

        keywords_to_save.append((keyword, reply, item_id))

    # 保存关键词（只保存文本关键词，保留图片关键词）
    try:
        success = db_manager.save_text_keywords_only(cid, keywords_to_save)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词失败")
    except Exception as e:
        error_msg = str(e)

        # 检查是否是图片关键词冲突
        if "已存在（图片关键词）" in error_msg:
            # 直接使用数据库管理器提供的友好错误信息
            raise HTTPException(status_code=400, detail=error_msg)
        elif "UNIQUE constraint failed" in error_msg or "唯一约束冲突" in error_msg:
            # 尝试从错误信息中提取具体的冲突关键词
            conflict_keyword = None
            conflict_type = None

            # 检查是否是数据库管理器抛出的详细错误
            if "关键词唯一约束冲突" in error_msg:
                # 解析详细错误信息：关键词唯一约束冲突: Cookie=xxx, 关键词='xxx', 通用关键词/商品ID: xxx
                import re
                keyword_match = re.search(r"关键词='([^']+)'", error_msg)
                if keyword_match:
                    conflict_keyword = keyword_match.group(1)

                if "通用关键词" in error_msg:
                    conflict_type = "通用关键词"
                elif "商品ID:" in error_msg:
                    item_match = re.search(r"商品ID: ([^\s,]+)", error_msg)
                    if item_match:
                        conflict_type = f"商品关键词（商品ID: {item_match.group(1)}）"

            # 构造用户友好的错误信息
            if conflict_keyword and conflict_type:
                detail_msg = f'关键词 "{conflict_keyword}" （{conflict_type}） 已存在，请使用其他关键词或商品ID'
            elif "keywords.cookie_id, keywords.keyword" in error_msg:
                detail_msg = "关键词重复！该关键词已存在（可能是图片关键词或文本关键词），请使用其他关键词"
            else:
                detail_msg = "关键词重复！请使用不同的关键词或商品ID组合"

            raise HTTPException(status_code=400, detail=detail_msg)
        else:
            log_with_user('error', f"保存关键词时发生未知错误: {error_msg}", current_user)
            raise HTTPException(status_code=500, detail="保存关键词失败")

    log_with_user('info', f"更新Cookie关键字(含商品ID): {cid}, 数量: {len(keywords_to_save)}", current_user)
    return {"msg": "updated", "count": len(keywords_to_save)}


@app.get("/items/{cid}")
def get_items_list(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的商品列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    try:
        # 获取该账号的所有商品
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            cursor.execute('''
            SELECT item_id, item_title, item_price, created_at
            FROM item_info
            WHERE cookie_id = ?
            ORDER BY created_at DESC
            ''', (cid,))

            items = []
            for row in cursor.fetchall():
                items.append({
                    'item_id': row[0],
                    'item_title': row[1] or '未知商品',
                    'item_price': row[2] or '价格未知',
                    'created_at': row[3]
                })

            return {"items": items, "count": len(items)}

    except Exception as e:
        logger.error(f"获取商品列表失败: {e}")
        raise HTTPException(status_code=500, detail="获取商品列表失败")


@app.get("/keywords-export/{cid}")
def export_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """导出指定账号的关键词为Excel文件"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    try:
        # 获取关键词数据（包含类型信息）
        keywords = db_manager.get_keywords_with_type(cid)

        # 创建DataFrame，只导出文本类型的关键词
        data = []
        for keyword_data in keywords:
            # 只导出文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                data.append({
                    '关键词': keyword_data['keyword'],
                    '商品ID': keyword_data['item_id'] or '',
                    '关键词内容': keyword_data['reply']
                })

        # 如果没有数据，创建空的DataFrame但保留列名（作为模板）
        if not data:
            df = pd.DataFrame(columns=['关键词', '商品ID', '关键词内容'])
        else:
            df = pd.DataFrame(data)

        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='关键词数据', index=False)

            # 如果是空模板，添加一些示例说明
            if data == []:
                worksheet = writer.sheets['关键词数据']
                # 添加示例数据作为注释（从第2行开始）
                worksheet['A2'] = '你好'
                worksheet['B2'] = ''
                worksheet['C2'] = '您好！欢迎咨询，有什么可以帮助您的吗？'

                worksheet['A3'] = '价格'
                worksheet['B3'] = '123456'
                worksheet['C3'] = '这个商品的价格是99元，现在有优惠活动哦！'

                worksheet['A4'] = '发货'
                worksheet['B4'] = ''
                worksheet['C4'] = '我们会在24小时内发货，请耐心等待。'

                # 设置示例行的样式（浅灰色背景）
                from openpyxl.styles import PatternFill
                gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                for row in range(2, 5):
                    for col in range(1, 4):
                        worksheet.cell(row=row, column=col).fill = gray_fill

        output.seek(0)

        # 生成文件名（使用URL编码处理中文）
        from urllib.parse import quote
        if not data:
            filename = f"keywords_template_{cid}_{int(time.time())}.xlsx"
        else:
            filename = f"keywords_{cid}_{int(time.time())}.xlsx"
        encoded_filename = quote(filename.encode('utf-8'))

        # 返回文件
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except Exception as e:
        logger.error(f"导出关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出关键词失败: {str(e)}")


@app.post("/keywords-import/{cid}")
async def import_keywords(cid: str, file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """导入Excel文件中的关键词到指定账号"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 检查文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx或.xls)")

    try:
        # 读取Excel文件
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))

        # 检查必要的列
        required_columns = ['关键词', '商品ID', '关键词内容']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Excel文件缺少必要的列: {', '.join(missing_columns)}")

        # 获取现有的文本类型关键词（用于比较更新/新增）
        existing_keywords = db_manager.get_keywords_with_type(cid)
        existing_dict = {}
        for keyword_data in existing_keywords:
            # 只考虑文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                keyword = keyword_data['keyword']
                reply = keyword_data['reply']
                item_id = keyword_data['item_id']
                key = f"{keyword}|{item_id or ''}"
                existing_dict[key] = (keyword, reply, item_id)

        # 处理导入数据
        import_data = []
        update_count = 0
        add_count = 0

        def clean_cell_value(value):
            """清理单元格值，处理数字转字符串时的 .0 后缀问题"""
            if pd.isna(value):
                return ''
            # 如果是数字类型，先转为整数（如果是整数值）再转字符串
            if isinstance(value, float) and value == int(value):
                return str(int(value)).strip()
            return str(value).strip()

        for index, row in df.iterrows():
            keyword = clean_cell_value(row['关键词'])
            item_id = clean_cell_value(row['商品ID']) or None
            reply = clean_cell_value(row['关键词内容'])

            if not keyword:
                continue  # 跳过没有关键词的行

            # 检查是否重复
            key = f"{keyword}|{item_id or ''}"
            if key in existing_dict:
                # 更新现有关键词
                update_count += 1
            else:
                # 新增关键词
                add_count += 1

            import_data.append((keyword, reply, item_id))

        if not import_data:
            raise HTTPException(status_code=400, detail="Excel文件中没有有效的关键词数据")

        # 保存到数据库（只影响文本关键词，保留图片关键词）
        success = db_manager.save_text_keywords_only(cid, import_data)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词到数据库失败")

        log_with_user('info', f"导入关键词成功: {cid}, 新增: {add_count}, 更新: {update_count}", current_user)

        return {
            "msg": "导入成功",
            "total": len(import_data),
            "added": add_count,
            "updated": update_count
        }

    except pd.errors.EmptyDataError:
        raise HTTPException(status_code=400, detail="Excel文件为空")
    except pd.errors.ParserError:
        raise HTTPException(status_code=400, detail="Excel文件格式错误")
    except Exception as e:
        logger.error(f"导入关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"导入关键词失败: {str(e)}")


@app.post("/keywords/{cid}/image")
async def add_image_keyword(
    cid: str,
    keyword: str = Form(...),
    item_id: str = Form(default=""),
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """添加图片关键词"""
    logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}")

    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查参数
    if not keyword or not keyword.strip():
        raise HTTPException(status_code=400, detail="关键词不能为空")

    if not image or not image.filename:
        raise HTTPException(status_code=400, detail="请选择图片文件")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}, filename={image.filename}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片保存成功: {image_url}")

        # 先检查关键词是否已存在
        normalized_item_id = item_id if item_id and item_id.strip() else None
        if db_manager.check_keyword_duplicate(cid, keyword, normalized_item_id):
            # 删除已保存的图片
            image_manager.delete_image(image_url)
            if normalized_item_id:
                raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' 在商品 '{normalized_item_id}' 中已存在")
            else:
                raise HTTPException(status_code=400, detail=f"通用关键词 '{keyword}' 已存在")

        # 保存图片关键词到数据库
        success = db_manager.save_image_keyword(cid, keyword, image_url, item_id or None)
        if not success:
            # 如果数据库保存失败，删除已保存的图片
            logger.error("数据库保存失败，删除已保存的图片")
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=400, detail="图片关键词保存失败，请稍后重试")

        log_with_user('info', f"添加图片关键词成功: {cid}, 关键词: {keyword}", current_user)

        return {
            "msg": "图片关键词添加成功",
            "keyword": keyword,
            "image_url": image_url,
            "item_id": item_id or None
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"添加图片关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"添加图片关键词失败: {str(e)}")


@app.post("/upload-image")
async def upload_image(
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """上传图片（用于卡券等功能）"""
    try:
        logger.info(f"接收到图片上传请求: filename={image.filename}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片上传成功: {image_url}")

        return {
            "message": "图片上传成功",
            "image_url": image_url
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"图片上传失败: {e}")
        raise HTTPException(status_code=500, detail=f"图片上传失败: {str(e)}")


@app.get("/keywords-with-type/{cid}")
def get_keywords_with_type(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含类型信息的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        keywords = db_manager.get_keywords_with_type(cid)
        return keywords
    except Exception as e:
        logger.error(f"获取关键词列表失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取关键词列表失败: {str(e)}")


@app.delete("/keywords/{cid}/{index}")
def delete_keyword_by_index(cid: str, index: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """根据索引删除关键词"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        # 先获取要删除的关键词信息（用于删除图片文件）
        keywords = db_manager.get_keywords_with_type(cid)
        if 0 <= index < len(keywords):
            keyword_data = keywords[index]

            # 删除关键词
            success = db_manager.delete_keyword_by_index(cid, index)
            if not success:
                raise HTTPException(status_code=400, detail="删除关键词失败")

            # 如果是图片关键词，删除对应的图片文件
            if keyword_data.get('type') == 'image' and keyword_data.get('image_url'):
                image_manager.delete_image(keyword_data['image_url'])

            log_with_user('info', f"删除关键词成功: {cid}, 索引: {index}, 关键词: {keyword_data.get('keyword')}", current_user)

            return {"msg": "删除成功"}
        else:
            raise HTTPException(status_code=400, detail="关键词索引无效")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"删除关键词失败: {str(e)}")


@app.get("/debug/keywords-table-info")
def debug_keywords_table_info(current_user: Dict[str, Any] = Depends(get_current_user)):
    """调试：检查keywords表结构"""
    try:
        import sqlite3
        conn = sqlite3.connect(db_manager.db_path)
        cursor = conn.cursor()

        # 获取表结构信息
        cursor.execute("PRAGMA table_info(keywords)")
        columns = cursor.fetchall()

        # 获取数据库版本
        cursor.execute("SELECT value FROM system_settings WHERE key = 'db_version'")
        version_result = cursor.fetchone()
        db_version = version_result[0] if version_result else "未知"

        conn.close()

        return {
            "db_version": db_version,
            "table_columns": [{"name": col[1], "type": col[2], "default": col[4]} for col in columns]
        }
    except Exception as e:
        logger.error(f"检查表结构失败: {e}")
        raise HTTPException(status_code=500, detail=f"检查表结构失败: {str(e)}")


# 卡券管理API
@app.get("/cards")
def get_cards(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的卡券列表"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        cards = db_manager.get_all_cards(user_id)
        return cards
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/cards")
def create_card(card_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建新卡券"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        card_name = card_data.get('name', '未命名卡券')

        log_with_user('info', f"创建卡券: {card_name}", current_user)

        # 验证多规格字段
        is_multi_spec = card_data.get('is_multi_spec', False)
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        card_id = db_manager.create_card(
            name=card_data.get('name'),
            card_type=card_data.get('type'),
            api_config=card_data.get('api_config'),
            text_content=card_data.get('text_content'),
            data_content=card_data.get('data_content'),
            image_url=card_data.get('image_url'),
            description=card_data.get('description'),
            enabled=card_data.get('enabled', True),
            delay_seconds=card_data.get('delay_seconds', 0),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name') if is_multi_spec else None,
            spec_value=card_data.get('spec_value') if is_multi_spec else None,
            user_id=user_id
        )

        log_with_user('info', f"卡券创建成功: {card_name} (ID: {card_id})", current_user)
        return {"id": card_id, "message": "卡券创建成功"}
    except Exception as e:
        log_with_user('error', f"创建卡券失败: {card_data.get('name', '未知')} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cards/{card_id}")
def get_card(card_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取单个卡券详情"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        card = db_manager.get_card_by_id(card_id, user_id)
        if card:
            return card
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cards/{card_id}")
def update_card(card_id: int, card_data: dict, _: None = Depends(require_auth)):
    """更新卡券"""
    try:
        from db_manager import db_manager
        # 验证多规格字段
        is_multi_spec = card_data.get('is_multi_spec')
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        success = db_manager.update_card(
            card_id=card_id,
            name=card_data.get('name'),
            card_type=card_data.get('type'),
            api_config=card_data.get('api_config'),
            text_content=card_data.get('text_content'),
            data_content=card_data.get('data_content'),
            image_url=card_data.get('image_url'),
            description=card_data.get('description'),
            enabled=card_data.get('enabled', True),
            delay_seconds=card_data.get('delay_seconds'),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name'),
            spec_value=card_data.get('spec_value')
        )
        if success:
            return {"message": "卡券更新成功"}
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cards/{card_id}/image")
async def update_card_with_image(
    card_id: int,
    image: UploadFile = File(...),
    name: str = Form(...),
    type: str = Form(...),
    description: str = Form(default=""),
    delay_seconds: int = Form(default=0),
    enabled: bool = Form(default=True),
    is_multi_spec: bool = Form(default=False),
    spec_name: str = Form(default=""),
    spec_value: str = Form(default=""),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新带图片的卡券"""
    try:
        logger.info(f"接收到带图片的卡券更新请求: card_id={card_id}, name={name}, type={type}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 验证多规格字段
        if is_multi_spec:
            if not spec_name or not spec_value:
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片保存成功: {image_url}")

        # 更新卡券
        from db_manager import db_manager
        success = db_manager.update_card(
            card_id=card_id,
            name=name,
            card_type=type,
            image_url=image_url,
            description=description,
            enabled=enabled,
            delay_seconds=delay_seconds,
            is_multi_spec=is_multi_spec,
            spec_name=spec_name if is_multi_spec else None,
            spec_value=spec_value if is_multi_spec else None
        )

        if success:
            logger.info(f"卡券更新成功: {name} (ID: {card_id})")
            return {"message": "卡券更新成功", "image_url": image_url}
        else:
            # 如果数据库更新失败，删除已保存的图片
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=404, detail="卡券不存在")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新带图片的卡券失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# 自动发货规则API
@app.get("/delivery-rules")
def get_delivery_rules(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取发货规则列表"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rules = db_manager.get_all_delivery_rules(user_id)
        return rules
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/delivery-rules")
def create_delivery_rule(rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建新发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rule_id = db_manager.create_delivery_rule(
            keyword=rule_data.get('keyword'),
            card_id=rule_data.get('card_id'),
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        return {"id": rule_id, "message": "发货规则创建成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/delivery-rules/{rule_id}")
def get_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取单个发货规则详情"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rule = db_manager.get_delivery_rule_by_id(rule_id, user_id)
        if rule:
            return rule
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/delivery-rules/{rule_id}")
def update_delivery_rule(rule_id: int, rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.update_delivery_rule(
            rule_id=rule_id,
            keyword=rule_data.get('keyword'),
            card_id=rule_data.get('card_id'),
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        if success:
            return {"message": "发货规则更新成功"}
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/cards/{card_id}")
def delete_card(card_id: int, _: None = Depends(require_auth)):
    """删除卡券"""
    try:
        from db_manager import db_manager
        success = db_manager.delete_card(card_id)
        if success:
            return {"message": "卡券删除成功"}
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/delivery-rules/{rule_id}")
def delete_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.delete_delivery_rule(rule_id, user_id)
        if success:
            return {"message": "发货规则删除成功"}
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 备份和恢复 API ====================

@app.get("/backup/export")
def export_backup(current_user: Dict[str, Any] = Depends(get_current_user)):
    """导出用户备份"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        username = current_user['username']

        # 导出当前用户的数据
        backup_data = db_manager.export_backup(user_id)

        # 生成文件名
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"xianyu_backup_{username}_{timestamp}.json"

        # 返回JSON响应，设置下载头
        response = JSONResponse(content=backup_data)
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/json"

        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出备份失败: {str(e)}")


@app.post("/backup/import")
def import_backup(file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """导入用户备份"""
    try:
        # 验证文件类型
        if not file.filename.endswith('.json'):
            raise HTTPException(status_code=400, detail="只支持JSON格式的备份文件")

        # 读取文件内容
        content = file.file.read()
        backup_data = json.loads(content.decode('utf-8'))

        # 导入备份到当前用户
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.import_backup(backup_data, user_id)

        if success:
            # 备份导入成功后，刷新 CookieManager 的内存缓存
            import cookie_manager
            if cookie_manager.manager:
                try:
                    cookie_manager.manager.reload_from_db()
                    logger.info("备份导入后已刷新 CookieManager 缓存")
                except Exception as e:
                    logger.error(f"刷新 CookieManager 缓存失败: {e}")

            return {"message": "备份导入成功"}
        else:
            raise HTTPException(status_code=400, detail="备份导入失败")

    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="备份文件格式无效")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入备份失败: {str(e)}")


@app.post("/system/reload-cache")
def reload_cache(_: None = Depends(require_auth)):
    """重新加载系统缓存（用于手动刷新数据）"""
    try:
        import cookie_manager
        if cookie_manager.manager:
            success = cookie_manager.manager.reload_from_db()
            if success:
                return {"message": "系统缓存已刷新", "success": True}
            else:
                raise HTTPException(status_code=500, detail="缓存刷新失败")
        else:
            raise HTTPException(status_code=500, detail="CookieManager 未初始化")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"刷新缓存失败: {str(e)}")


# ==================== 商品管理 API ====================

@app.get("/items")
def get_all_items(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的所有商品信息"""
    try:
        # 只返回当前用户的商品信息
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_items_by_cookie(cookie_id)
            all_items.extend(items)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")


# ==================== 商品搜索 API ====================

class ItemSearchRequest(BaseModel):
    keyword: str
    page: int = 1
    page_size: int = 20

class ItemSearchMultipleRequest(BaseModel):
    keyword: str
    total_pages: int = 1

@app.post("/items/search")
async def search_items(
    search_request: ItemSearchRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """搜索闲鱼商品"""
    user_info = f"【{current_user.get('username', 'unknown')}#{current_user.get('user_id', 'unknown')}】" if current_user else "【未登录】"

    try:
        logger.info(f"{user_info} 开始单页搜索: 关键词='{search_request.keyword}', 页码={search_request.page}, 每页={search_request.page_size}")

        from utils.item_search import search_xianyu_items

        # 执行搜索
        result = await search_xianyu_items(
            keyword=search_request.keyword,
            page=search_request.page,
            page_size=search_request.page_size
        )

        # 检查是否有错误
        has_error = result.get("error")
        items_count = len(result.get("items", []))

        logger.info(f"{user_info} 单页搜索完成: 获取到 {items_count} 条数据" +
                   (f", 错误: {has_error}" if has_error else ""))

        response_data = {
            "success": True,
            "data": result.get("items", []),
            "total": result.get("total", 0),
            "page": search_request.page,
            "page_size": search_request.page_size,
            "keyword": search_request.keyword,
            "is_real_data": result.get("is_real_data", False),
            "source": result.get("source", "unknown")
        }

        # 如果有错误信息，也包含在响应中
        if has_error:
            response_data["error"] = has_error

        return response_data

    except Exception as e:
        error_msg = str(e)
        logger.error(f"{user_info} 商品搜索失败: {error_msg}")
        raise HTTPException(status_code=500, detail=f"商品搜索失败: {error_msg}")


@app.get("/cookies/check")
async def check_valid_cookies(
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """检查是否有有效的cookies账户（必须是启用状态）"""
    try:
        if cookie_manager.manager is None:
            return {
                "success": True,
                "hasValidCookies": False,
                "validCount": 0,
                "enabledCount": 0,
                "totalCount": 0
            }

        from db_manager import db_manager

        # 获取所有cookies
        all_cookies = db_manager.get_all_cookies()

        # 检查启用状态和有效性
        valid_cookies = []
        enabled_cookies = []

        for cookie_id, cookie_value in all_cookies.items():
            # 检查是否启用
            is_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
            if is_enabled:
                enabled_cookies.append(cookie_id)
                # 检查是否有效（长度大于50）
                if len(cookie_value) > 50:
                    valid_cookies.append(cookie_id)

        return {
            "success": True,
            "hasValidCookies": len(valid_cookies) > 0,
            "validCount": len(valid_cookies),
            "enabledCount": len(enabled_cookies),
            "totalCount": len(all_cookies)
        }

    except Exception as e:
        logger.error(f"检查cookies失败: {str(e)}")
        return {
            "success": False,
            "hasValidCookies": False,
            "error": str(e)
        }

@app.post("/items/search_multiple")
async def search_multiple_pages(
    search_request: ItemSearchMultipleRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """搜索多页闲鱼商品"""
    user_info = f"【{current_user.get('username', 'unknown')}#{current_user.get('user_id', 'unknown')}】" if current_user else "【未登录】"

    try:
        logger.info(f"{user_info} 开始多页搜索: 关键词='{search_request.keyword}', 页数={search_request.total_pages}")

        from utils.item_search import search_multiple_pages_xianyu

        # 执行多页搜索
        result = await search_multiple_pages_xianyu(
            keyword=search_request.keyword,
            total_pages=search_request.total_pages
        )

        # 检查是否有错误
        has_error = result.get("error")
        items_count = len(result.get("items", []))

        logger.info(f"{user_info} 多页搜索完成: 获取到 {items_count} 条数据" +
                   (f", 错误: {has_error}" if has_error else ""))

        response_data = {
            "success": True,
            "data": result.get("items", []),
            "total": result.get("total", 0),
            "total_pages": search_request.total_pages,
            "keyword": search_request.keyword,
            "is_real_data": result.get("is_real_data", False),
            "is_fallback": result.get("is_fallback", False),
            "source": result.get("source", "unknown")
        }

        # 如果有错误信息，也包含在响应中
        if has_error:
            response_data["error"] = has_error

        return response_data

    except Exception as e:
        error_msg = str(e)
        logger.error(f"{user_info} 多页商品搜索失败: {error_msg}")
        raise HTTPException(status_code=500, detail=f"多页商品搜索失败: {error_msg}")



@app.get("/items/cookie/{cookie_id}")
def get_items_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定Cookie的商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        items = db_manager.get_items_by_cookie(cookie_id)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")


@app.get("/items/{cookie_id}/{item_id}")
def get_item_detail(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取商品详情"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        item = db_manager.get_item_info(cookie_id, item_id)
        if not item:
            raise HTTPException(status_code=404, detail="商品不存在")
        return {"item": item}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品详情失败: {str(e)}")


class ItemDetailUpdate(BaseModel):
    item_detail: str


@app.put("/items/{cookie_id}/{item_id}")
def update_item_detail(
    cookie_id: str,
    item_id: str,
    update_data: ItemDetailUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新商品详情"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.update_item_detail(cookie_id, item_id, update_data.item_detail)
        if success:
            return {"message": "商品详情更新成功"}
        else:
            raise HTTPException(status_code=400, detail="更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新商品详情失败: {str(e)}")


@app.delete("/items/{cookie_id}/{item_id}")
def delete_item_info(
    cookie_id: str,
    item_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """删除商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.delete_item_info(cookie_id, item_id)
        if success:
            return {"message": "商品信息删除成功"}
        else:
            raise HTTPException(status_code=404, detail="商品信息不存在")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除商品信息异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


class BatchDeleteRequest(BaseModel):
    items: List[dict]  # [{"cookie_id": "xxx", "item_id": "yyy"}, ...]


class AIReplySettings(BaseModel):
    ai_enabled: bool
    model_name: str = "qwen-plus"
    api_key: str = ""
    base_url: str = "https://dashscope.aliyuncs.com/compatible-mode/v1"
    max_discount_percent: int = 10
    max_discount_amount: int = 100
    max_bargain_rounds: int = 3
    custom_prompts: str = ""


@app.delete("/items/batch")
def batch_delete_items(
    request: BatchDeleteRequest,
    _: None = Depends(require_auth)
):
    """批量删除商品信息"""
    try:
        if not request.items:
            raise HTTPException(status_code=400, detail="删除列表不能为空")

        success_count = db_manager.batch_delete_item_info(request.items)
        total_count = len(request.items)

        return {
            "message": f"批量删除完成",
            "success_count": success_count,
            "total_count": total_count,
            "failed_count": total_count - success_count
        }
    except Exception as e:
        logger.error(f"批量删除商品信息异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


# ==================== AI回复管理API ====================

@app.get("/ai-reply-settings/{cookie_id}")
def get_ai_reply_settings(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的AI回复设置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        settings = db_manager.get_ai_reply_settings(cookie_id)
        return settings
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.put("/ai-reply-settings/{cookie_id}")
def update_ai_reply_settings(cookie_id: str, settings: AIReplySettings, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的AI回复设置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 检查账号是否存在
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager 未就绪')

        # 保存设置
        settings_dict = settings.dict()
        success = db_manager.save_ai_reply_settings(cookie_id, settings_dict)

        if success:

            # 如果启用了AI回复，记录日志
            if settings.ai_enabled:
                logger.info(f"账号 {cookie_id} 启用AI回复")
            else:
                logger.info(f"账号 {cookie_id} 禁用AI回复")

            return {"message": "AI回复设置更新成功"}
        else:
            raise HTTPException(status_code=400, detail="更新失败")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.get("/ai-reply-settings")
def get_all_ai_reply_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的AI回复设置"""
    try:
        # 只返回当前用户的AI回复设置
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_settings = db_manager.get_all_ai_reply_settings()
        # 过滤只属于当前用户的设置
        user_settings = {cid: settings for cid, settings in all_settings.items() if cid in user_cookies}
        return user_settings
    except Exception as e:
        logger.error(f"获取所有AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.post("/ai-reply-test/{cookie_id}")
def test_ai_reply(cookie_id: str, test_data: dict, _: None = Depends(require_auth)):
    """测试AI回复功能"""
    try:
        # 检查账号是否存在
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager 未就绪')

        if cookie_id not in cookie_manager.manager.cookies:
            raise HTTPException(status_code=404, detail='账号不存在')

        # 检查是否启用AI回复
        if not ai_reply_engine.is_ai_enabled(cookie_id):
            raise HTTPException(status_code=400, detail='该账号未启用AI回复')

        # 检查AI设置是否完整
        settings = db_manager.get_ai_reply_settings(cookie_id)
        if not settings.get('api_key'):
            raise HTTPException(status_code=400, detail='未配置API Key，请先在AI设置中配置API Key')
        if not settings.get('base_url'):
            raise HTTPException(status_code=400, detail='未配置API地址，请先在AI设置中配置API地址')

        # 构造测试数据
        test_message = test_data.get('message', '你好')
        test_item_info = {
            'title': test_data.get('item_title', '测试商品'),
            'price': test_data.get('item_price', 100),
            'desc': test_data.get('item_desc', '这是一个测试商品')
        }

        # 生成测试回复（跳过等待时间）
        reply = ai_reply_engine.generate_reply(
            message=test_message,
            item_info=test_item_info,
            chat_id=f"test_{int(time.time())}",
            cookie_id=cookie_id,
            user_id="test_user",
            item_id="test_item",
            skip_wait=True  # 测试时跳过10秒等待
        )

        if reply:
            return {"message": "测试成功", "reply": reply}
        else:
            raise HTTPException(status_code=400, detail="AI回复生成失败，请检查API Key是否正确、API地址是否可访问")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"测试AI回复异常: {e}")
        import traceback
        logger.error(f"详细错误: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


# ==================== 日志管理API ====================

@app.get("/logs")
async def get_logs(lines: int = 200, level: str = None, source: str = None, _: None = Depends(require_auth)):
    """获取实时系统日志"""
    try:
        # 获取文件日志收集器
        collector = get_file_log_collector()

        # 获取日志
        logs = collector.get_logs(lines=lines, level_filter=level, source_filter=source)

        return {"success": True, "logs": logs}

    except Exception as e:
        return {"success": False, "message": f"获取日志失败: {str(e)}", "logs": []}


@app.get("/risk-control-logs")
async def get_risk_control_logs(
    cookie_id: str = None,
    limit: int = 100,
    offset: int = 0,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """获取风控日志（管理员专用）"""
    try:
        log_with_user('info', f"查询风控日志: cookie_id={cookie_id}, limit={limit}, offset={offset}", admin_user)

        # 获取风控日志
        logs = db_manager.get_risk_control_logs(cookie_id=cookie_id, limit=limit, offset=offset)
        total_count = db_manager.get_risk_control_logs_count(cookie_id=cookie_id)

        log_with_user('info', f"风控日志查询成功，共 {len(logs)} 条记录，总计 {total_count} 条", admin_user)

        return {
            "success": True,
            "data": logs,
            "total": total_count,
            "limit": limit,
            "offset": offset
        }

    except Exception as e:
        log_with_user('error', f"获取风控日志失败: {str(e)}", admin_user)
        return {
            "success": False,
            "message": f"获取风控日志失败: {str(e)}",
            "data": [],
            "total": 0
        }


@app.delete("/risk-control-logs/{log_id}")
async def delete_risk_control_log(
    log_id: int,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """删除风控日志记录（管理员专用）"""
    try:
        log_with_user('info', f"删除风控日志记录: {log_id}", admin_user)

        success = db_manager.delete_risk_control_log(log_id)

        if success:
            log_with_user('info', f"风控日志删除成功: {log_id}", admin_user)
            return {"success": True, "message": "删除成功"}
        else:
            log_with_user('warning', f"风控日志删除失败: {log_id}", admin_user)
            return {"success": False, "message": "删除失败，记录可能不存在"}

    except Exception as e:
        log_with_user('error', f"删除风控日志失败: {log_id} - {str(e)}", admin_user)
        return {"success": False, "message": f"删除失败: {str(e)}"}


@app.get("/logs/stats")
async def get_log_stats(_: None = Depends(require_auth)):
    """获取日志统计信息"""
    try:
        collector = get_file_log_collector()
        stats = collector.get_stats()

        return {"success": True, "stats": stats}

    except Exception as e:
        return {"success": False, "message": f"获取日志统计失败: {str(e)}", "stats": {}}


@app.post("/logs/clear")
async def clear_logs(_: None = Depends(require_auth)):
    """清空日志"""
    try:
        collector = get_file_log_collector()
        collector.clear_logs()

        return {"success": True, "message": "日志已清空"}

    except Exception as e:
        return {"success": False, "message": f"清空日志失败: {str(e)}"}


# ==================== 商品管理API ====================

@app.post("/items/get-all-from-account")
async def get_all_items_from_account(request: dict, _: None = Depends(require_auth)):
    """从指定账号获取所有商品信息"""
    try:
        cookie_id = request.get('cookie_id')
        if not cookie_id:
            return {"success": False, "message": "缺少cookie_id参数"}

        # 获取指定账号的cookie信息
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {"success": False, "message": "未找到指定的账号信息"}

        cookies_str = cookie_info.get('cookies_str', '')
        if not cookies_str:
            return {"success": False, "message": "账号cookie信息为空"}

        # 创建XianyuLive实例，传入正确的cookie_id
        from XianyuAutoAsync import XianyuLive
        xianyu_instance = XianyuLive(cookies_str, cookie_id)

        # 调用获取所有商品信息的方法（自动分页）
        logger.info(f"开始获取账号 {cookie_id} 的所有商品信息")
        result = await xianyu_instance.get_all_items()

        # 关闭session
        await xianyu_instance.close_session()

        if result.get('error'):
            logger.error(f"获取商品信息失败: {result['error']}")
            return {"success": False, "message": result['error']}
        else:
            total_count = result.get('total_count', 0)
            total_pages = result.get('total_pages', 1)
            saved_count = result.get('total_saved', 0)
            logger.info(f"成功获取账号 {cookie_id} 的 {total_count} 个商品（共{total_pages}页），保存 {saved_count} 个")
            return {
                "success": True,
                "message": f"成功获取商品，共 {total_count} 件，保存 {saved_count} 件",
                "total_count": total_count,
                "total_pages": total_pages,
                "saved_count": saved_count
            }

    except Exception as e:
        logger.error(f"获取账号商品信息异常: {str(e)}")
        return {"success": False, "message": f"获取商品信息异常: {str(e)}"}


@app.post("/items/get-by-page")
async def get_items_by_page(request: dict, _: None = Depends(require_auth)):
    """从指定账号按页获取商品信息"""
    try:
        # 验证参数
        cookie_id = request.get('cookie_id')
        page_number = request.get('page_number', 1)
        page_size = request.get('page_size', 20)

        if not cookie_id:
            return {"success": False, "message": "缺少cookie_id参数"}

        # 验证分页参数
        try:
            page_number = int(page_number)
            page_size = int(page_size)
        except (ValueError, TypeError):
            return {"success": False, "message": "页码和每页数量必须是数字"}

        if page_number < 1:
            return {"success": False, "message": "页码必须大于0"}

        if page_size < 1 or page_size > 100:
            return {"success": False, "message": "每页数量必须在1-100之间"}

        # 获取账号信息
        account = db_manager.get_cookie_by_id(cookie_id)
        if not account:
            return {"success": False, "message": "账号不存在"}

        cookies_str = account['cookies_str']
        if not cookies_str:
            return {"success": False, "message": "账号cookies为空"}

        # 创建XianyuLive实例，传入正确的cookie_id
        from XianyuAutoAsync import XianyuLive
        xianyu_instance = XianyuLive(cookies_str, cookie_id)

        # 调用获取指定页商品信息的方法
        logger.info(f"开始获取账号 {cookie_id} 第{page_number}页商品信息（每页{page_size}条）")
        result = await xianyu_instance.get_item_list_info(page_number, page_size)

        # 关闭session
        await xianyu_instance.close_session()

        if result.get('error'):
            logger.error(f"获取商品信息失败: {result['error']}")
            return {"success": False, "message": result['error']}
        else:
            current_count = result.get('current_count', 0)
            logger.info(f"成功获取账号 {cookie_id} 第{page_number}页 {current_count} 个商品")
            return {
                "success": True,
                "message": f"成功获取第{page_number}页 {current_count} 个商品，详细信息已打印到控制台",
                "page_number": page_number,
                "page_size": page_size,
                "current_count": current_count
            }

    except Exception as e:
        logger.error(f"获取账号商品信息异常: {str(e)}")
        return {"success": False, "message": f"获取商品信息异常: {str(e)}"}


# ------------------------- 用户设置接口 -------------------------

@app.get('/user-settings')
def get_user_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        settings = db_manager.get_user_settings(user_id)
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put('/user-settings/{key}')
def update_user_setting(key: str, setting_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新用户设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        value = setting_data.get('value')
        description = setting_data.get('description', '')

        log_with_user('info', f"更新用户设置: {key} = {value}", current_user)

        success = db_manager.set_user_setting(user_id, key, value, description)
        if success:
            log_with_user('info', f"用户设置更新成功: {key}", current_user)
            return {'msg': 'setting updated', 'key': key, 'value': value}
        else:
            log_with_user('error', f"用户设置更新失败: {key}", current_user)
            raise HTTPException(status_code=400, detail='更新失败')
    except Exception as e:
        log_with_user('error', f"更新用户设置异常: {key} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/user-settings/{key}')
def get_user_setting(key: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取用户特定设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        setting = db_manager.get_user_setting(user_id, key)
        if setting:
            return setting
        else:
            raise HTTPException(status_code=404, detail='设置不存在')
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 管理员专用接口 -------------------------

@app.get('/admin/users')
def get_all_users(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取所有用户信息（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', "查询所有用户信息", admin_user)
        users = db_manager.get_all_users()

        # 为每个用户添加统计信息
        for user in users:
            user_id = user['id']
            # 统计用户的Cookie数量
            user_cookies = db_manager.get_all_cookies(user_id)
            user['cookie_count'] = len(user_cookies)

            # 统计用户的卡券数量
            user_cards = db_manager.get_all_cards(user_id)
            user['card_count'] = len(user_cards) if user_cards else 0

            # 隐藏密码字段
            if 'password_hash' in user:
                del user['password_hash']

        log_with_user('info', f"返回用户信息，共 {len(users)} 个用户", admin_user)
        return {"users": users}
    except Exception as e:
        log_with_user('error', f"获取用户信息失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/users/{user_id}')
def delete_user(user_id: int, admin_user: Dict[str, Any] = Depends(require_admin)):
    """删除用户（管理员专用）"""
    from db_manager import db_manager
    try:
        # 不能删除管理员自己
        if user_id == admin_user['user_id']:
            log_with_user('warning', "尝试删除管理员自己", admin_user)
            raise HTTPException(status_code=400, detail="不能删除管理员自己")

        # 获取要删除的用户信息
        user_to_delete = db_manager.get_user_by_id(user_id)
        if not user_to_delete:
            raise HTTPException(status_code=404, detail="用户不存在")

        log_with_user('info', f"准备删除用户: {user_to_delete['username']} (ID: {user_id})", admin_user)

        # 删除用户及其相关数据
        success = db_manager.delete_user_and_data(user_id)

        if success:
            log_with_user('info', f"用户删除成功: {user_to_delete['username']} (ID: {user_id})", admin_user)
            return {"message": f"用户 {user_to_delete['username']} 删除成功"}
        else:
            log_with_user('error', f"用户删除失败: {user_to_delete['username']} (ID: {user_id})", admin_user)
            raise HTTPException(status_code=400, detail="删除失败")
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除用户异常: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/risk-control-logs')
async def get_admin_risk_control_logs(
    cookie_id: str = None,
    limit: int = 100,
    offset: int = 0,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """获取风控日志（管理员专用）"""
    try:
        log_with_user('info', f"查询风控日志: cookie_id={cookie_id}, limit={limit}, offset={offset}", admin_user)

        # 获取风控日志
        logs = db_manager.get_risk_control_logs(cookie_id=cookie_id, limit=limit, offset=offset)
        total_count = db_manager.get_risk_control_logs_count(cookie_id=cookie_id)

        log_with_user('info', f"风控日志查询成功，共 {len(logs)} 条记录，总计 {total_count} 条", admin_user)

        return {
            "success": True,
            "data": logs,
            "total": total_count,
            "limit": limit,
            "offset": offset
        }

    except Exception as e:
        log_with_user('error', f"查询风控日志失败: {str(e)}", admin_user)
        return {"success": False, "message": f"查询失败: {str(e)}", "data": [], "total": 0}


@app.get('/admin/cookies')
def get_admin_cookies(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取所有Cookie信息（管理员专用）"""
    try:
        log_with_user('info', "查询所有Cookie信息", admin_user)

        if cookie_manager.manager is None:
            return {
                "success": True,
                "cookies": [],
                "message": "CookieManager 未就绪"
            }

        # 获取所有用户的cookies
        from db_manager import db_manager
        all_users = db_manager.get_all_users()
        all_cookies = []

        for user in all_users:
            user_id = user['id']
            user_cookies = db_manager.get_all_cookies(user_id)
            for cookie_id, cookie_value in user_cookies.items():
                # 获取cookie详细信息
                cookie_details = db_manager.get_cookie_details(cookie_id)
                cookie_info = {
                    'cookie_id': cookie_id,
                    'user_id': user_id,
                    'username': user['username'],
                    'nickname': cookie_details.get('remark', '') if cookie_details else '',
                    'enabled': cookie_manager.manager.get_cookie_status(cookie_id)
                }
                all_cookies.append(cookie_info)

        log_with_user('info', f"获取到 {len(all_cookies)} 个Cookie", admin_user)
        return {
            "success": True,
            "cookies": all_cookies,
            "total": len(all_cookies)
        }

    except Exception as e:
        log_with_user('error', f"获取Cookie信息失败: {str(e)}", admin_user)
        return {
            "success": False,
            "cookies": [],
            "message": f"获取失败: {str(e)}"
        }


@app.get('/admin/logs')
def get_system_logs(admin_user: Dict[str, Any] = Depends(require_admin),
                   lines: int = 100,
                   level: str = None):
    """获取系统日志（管理员专用）"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', f"查询系统日志，行数: {lines}, 级别: {level}", admin_user)

        # 查找日志文件
        log_files = glob.glob("logs/xianyu_*.log")
        logger.info(f"找到日志文件: {log_files}")

        if not log_files:
            logger.warning("未找到日志文件")
            return {"logs": [], "message": "未找到日志文件", "success": False}

        # 获取最新的日志文件
        latest_log_file = max(log_files, key=os.path.getctime)
        logger.info(f"使用最新日志文件: {latest_log_file}")

        logs = []
        try:
            with open(latest_log_file, 'r', encoding='utf-8') as f:
                all_lines = f.readlines()
                logger.info(f"读取到 {len(all_lines)} 行日志")

                # 如果指定了日志级别，进行过滤
                if level:
                    filtered_lines = [line for line in all_lines if f"| {level.upper()} |" in line]
                    logger.info(f"按级别 {level} 过滤后剩余 {len(filtered_lines)} 行")
                else:
                    filtered_lines = all_lines

                # 获取最后N行
                recent_lines = filtered_lines[-lines:] if len(filtered_lines) > lines else filtered_lines
                logger.info(f"取最后 {len(recent_lines)} 行日志")

                for line in recent_lines:
                    logs.append(line.strip())

        except Exception as e:
            logger.error(f"读取日志文件失败: {str(e)}")
            log_with_user('error', f"读取日志文件失败: {str(e)}", admin_user)
            return {"logs": [], "message": f"读取日志文件失败: {str(e)}", "success": False}

        log_with_user('info', f"返回日志记录 {len(logs)} 条", admin_user)
        logger.info(f"成功返回 {len(logs)} 条日志记录")

        return {
            "logs": logs,
            "log_file": latest_log_file,
            "total_lines": len(logs),
            "success": True
        }

    except Exception as e:
        logger.error(f"获取系统日志失败: {str(e)}")
        log_with_user('error', f"获取系统日志失败: {str(e)}", admin_user)
        return {"logs": [], "message": f"获取系统日志失败: {str(e)}", "success": False}

@app.get('/admin/log-files')
def list_log_files(admin_user: Dict[str, Any] = Depends(require_admin)):
    """列出所有可用的系统日志文件"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', "查询日志文件列表", admin_user)

        log_dir = "logs"
        if not os.path.exists(log_dir):
            logger.warning("日志目录不存在")
            return {"success": True, "files": []}

        log_pattern = os.path.join(log_dir, "xianyu_*.log")
        log_files = glob.glob(log_pattern)

        files_info = []
        for file_path in log_files:
            try:
                stat_info = os.stat(file_path)
                files_info.append({
                    "name": os.path.basename(file_path),
                    "size": stat_info.st_size,
                    "modified_at": datetime.fromtimestamp(stat_info.st_mtime).isoformat(),
                    "modified_ts": stat_info.st_mtime
                })
            except OSError as e:
                logger.warning(f"读取日志文件信息失败 {file_path}: {e}")

        # 按修改时间倒序排序
        files_info.sort(key=lambda item: item.get("modified_ts", 0), reverse=True)

        logger.info(f"返回日志文件列表，共 {len(files_info)} 个文件")
        return {"success": True, "files": files_info}

    except Exception as e:
        logger.error(f"获取日志文件列表失败: {str(e)}")
        log_with_user('error', f"获取日志文件列表失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/logs/export')
def export_log_file(file: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """导出指定的日志文件"""
    import os
    from fastapi.responses import StreamingResponse

    try:
        if not file:
            raise HTTPException(status_code=400, detail="缺少文件参数")

        safe_name = os.path.basename(file)
        log_dir = os.path.abspath("logs")
        target_path = os.path.abspath(os.path.join(log_dir, safe_name))

        # 防止目录遍历
        if not target_path.startswith(log_dir):
            log_with_user('warning', f"尝试访问非法日志文件: {file}", admin_user)
            raise HTTPException(status_code=400, detail="非法的日志文件路径")

        if not os.path.exists(target_path):
            log_with_user('warning', f"日志文件不存在: {file}", admin_user)
            raise HTTPException(status_code=404, detail="日志文件不存在")

        log_with_user('info', f"导出日志文件: {safe_name}", admin_user)
        def iter_file(path: str):
            file_handle = open(path, 'rb')
            try:
                while True:
                    chunk = file_handle.read(8192)
                    if not chunk:
                        break
                    yield chunk
            finally:
                file_handle.close()

        headers = {
            "Content-Disposition": f'attachment; filename="{safe_name}"'
        }
        return StreamingResponse(
            iter_file(target_path),
            media_type='text/plain; charset=utf-8',
            headers=headers
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出日志文件失败: {str(e)}")
        log_with_user('error', f"导出日志文件失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/stats')
def get_system_stats(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取系统统计信息（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', "查询系统统计信息", admin_user)

        # 用户统计
        all_users = db_manager.get_all_users()
        total_users = len(all_users)

        # Cookie统计
        all_cookies = db_manager.get_all_cookies()
        total_cookies = len(all_cookies)
        
        # 活跃账号统计（启用状态的账号）
        active_cookies = 0
        for cookie_id in all_cookies.keys():
            status = db_manager.get_cookie_status(cookie_id)
            if status:
                active_cookies += 1

        # 卡券统计
        all_cards = db_manager.get_all_cards()
        total_cards = len(all_cards) if all_cards else 0

        # 关键词统计
        all_keywords = db_manager.get_all_keywords()
        total_keywords = sum(len(kw_list) for kw_list in all_keywords.values())

        # 订单统计
        total_orders = 0
        try:
            orders = db_manager.get_all_orders()
            total_orders = len(orders) if orders else 0
        except:
            pass

        stats = {
            "total_users": total_users,
            "total_cookies": total_cookies,
            "active_cookies": active_cookies,
            "total_cards": total_cards,
            "total_keywords": total_keywords,
            "total_orders": total_orders
        }

        log_with_user('info', f"系统统计信息查询完成: {stats}", admin_user)
        return stats

    except Exception as e:
        log_with_user('error', f"获取系统统计信息失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

# ------------------------- BI报表分析接口 -------------------------

@app.get('/analytics/orders')
def get_order_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    获取订单分析数据（BI报表）

    Args:
        start_date: 开始日期 (格式: YYYY-MM-DD)
        end_date: 结束日期 (格式: YYYY-MM-DD)
    """
    from db_manager import db_manager
    try:
        log_with_user('info', f"查询订单分析数据: {start_date} - {end_date}", current_user)

        # 获取当前用户的ID
        user_id = current_user['user_id']

        # 定义有效订单状态（只统计这几种状态）
        valid_statuses = ['pending_ship', 'shipped', 'completed']

        # 调用数据库分析函数，传入包含状态
        analytics_data = db_manager.get_order_analytics(
            start_date=start_date,
            end_date=end_date,
            user_id=user_id,
            include_statuses=valid_statuses
        )

        if 'error' in analytics_data:
            log_with_user('error', f"获取订单分析数据失败: {analytics_data['error']}", current_user)
            raise HTTPException(status_code=500, detail=analytics_data['error'])

        log_with_user('info', "订单分析数据查询成功", current_user)
        return analytics_data

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"获取订单分析数据失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/analytics/orders/valid')
def get_valid_orders(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    获取有效订单详情列表（用于统计中的订单明细）

    Args:
        start_date: 开始日期 (格式: YYYY-MM-DD)
        end_date: 结束日期 (格式: YYYY-MM-DD)
    """
    from db_manager import db_manager
    try:
        log_with_user('info', f"查询有效订单列表: {start_date} - {end_date}", current_user)

        # 获取当前用户的ID
        user_id = current_user['user_id']

        # 定义有效订单状态
        valid_statuses = ['pending_ship', 'shipped', 'completed']

        # 调用数据库函数获取有效订单
        orders = db_manager.get_orders_for_analytics(
            start_date=start_date,
            end_date=end_date,
            user_id=user_id,
            include_statuses=valid_statuses
        )

        log_with_user('info', f"查询到 {len(orders)} 个有效订单", current_user)
        return {"orders": orders}

    except Exception as e:
        log_with_user('error', f"获取有效订单列表失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))

# ------------------------- 指定商品回复接口 -------------------------

@app.get("/itemReplays")
def get_all_items(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的所有商品回复信息"""
    try:
        # 只返回当前用户的商品信息
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_itemReplays_by_cookie(cookie_id)
            all_items.extend(items)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品回复信息失败: {str(e)}")

@app.get("/itemReplays/cookie/{cookie_id}")
def get_items_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定Cookie的商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        items = db_manager.get_itemReplays_by_cookie(cookie_id)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")

@app.put("/item-reply/{cookie_id}/{item_id}")
def update_item_reply(
    cookie_id: str,
    item_id: str,
    data: dict,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    更新指定账号和商品的回复内容
    """
    try:
        user_id = current_user['user_id']
        from db_manager import db_manager

        # 验证cookie是否属于用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        reply_content = data.get("reply_content", "").strip()
        if not reply_content:
            raise HTTPException(status_code=400, detail="回复内容不能为空")

        db_manager.update_item_reply(cookie_id=cookie_id, item_id=item_id, reply_content=reply_content)

        return {"message": "商品回复更新成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新商品回复失败: {str(e)}")

@app.delete("/item-reply/{cookie_id}/{item_id}")
def delete_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    删除指定账号cookie_id和商品item_id的商品回复
    """
    try:
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        success = db_manager.delete_item_reply(cookie_id, item_id)
        if not success:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return {"message": "商品回复删除成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除商品回复失败: {str(e)}")

class ItemToDelete(BaseModel):
    cookie_id: str
    item_id: str

class BatchDeleteRequest(BaseModel):
    items: List[ItemToDelete]

@app.delete("/item-reply/batch")
async def batch_delete_item_reply(
    req: BatchDeleteRequest,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    批量删除商品回复
    """
    user_id = current_user['user_id']
    from db_manager import db_manager

    # 先校验当前用户是否有权限删除每个cookie对应的回复
    user_cookies = db_manager.get_all_cookies(user_id)
    for item in req.items:
        if item.cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail=f"无权限访问Cookie {item.cookie_id}")

    result = db_manager.batch_delete_item_replies([item.dict() for item in req.items])
    return {
        "success_count": result["success_count"],
        "failed_count": result["failed_count"]
    }

@app.get("/item-reply/{cookie_id}/{item_id}")
def get_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    获取指定账号cookie_id和商品item_id的商品回复内容
    """
    try:
        user_id = current_user['user_id']
        # 校验cookie_id是否属于当前用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        # 获取指定商品回复
        item_replies = db_manager.get_itemReplays_by_cookie(cookie_id)
        # 找对应item_id的回复
        item_reply = next((r for r in item_replies if r['item_id'] == item_id), None)

        if item_reply is None:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return item_reply

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品回复失败: {str(e)}")


# ------------------------- 数据库备份和恢复接口 -------------------------

@app.get('/admin/backup/download')
def download_database_backup(admin_user: Dict[str, Any] = Depends(require_admin)):
    """下载数据库备份文件（管理员专用）"""
    import os
    from fastapi.responses import FileResponse
    from datetime import datetime

    try:
        log_with_user('info', "请求下载数据库备份", admin_user)

        # 使用db_manager的实际数据库路径
        from db_manager import db_manager
        db_file_path = db_manager.db_path

        # 检查数据库文件是否存在
        if not os.path.exists(db_file_path):
            log_with_user('error', f"数据库文件不存在: {db_file_path}", admin_user)
            raise HTTPException(status_code=404, detail="数据库文件不存在")

        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"xianyu_backup_{timestamp}.db"

        log_with_user('info', f"开始下载数据库备份: {download_filename}", admin_user)

        return FileResponse(
            path=db_file_path,
            filename=download_filename,
            media_type='application/octet-stream'
        )

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"下载数据库备份失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.post('/admin/backup/upload')
async def upload_database_backup(admin_user: Dict[str, Any] = Depends(require_admin),
                                backup_file: UploadFile = File(...)):
    """上传并恢复数据库备份文件（管理员专用）"""
    import os
    import shutil
    import sqlite3
    from datetime import datetime

    try:
        log_with_user('info', f"开始上传数据库备份: {backup_file.filename}", admin_user)

        # 验证文件类型
        if not backup_file.filename.endswith('.db'):
            log_with_user('warning', f"无效的备份文件类型: {backup_file.filename}", admin_user)
            raise HTTPException(status_code=400, detail="只支持.db格式的数据库文件")

        # 验证文件大小（限制100MB）
        content = await backup_file.read()
        if len(content) > 100 * 1024 * 1024:  # 100MB
            log_with_user('warning', f"备份文件过大: {len(content)} bytes", admin_user)
            raise HTTPException(status_code=400, detail="备份文件大小不能超过100MB")

        # 验证是否为有效的SQLite数据库文件
        temp_file_path = f"temp_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"

        try:
            # 保存临时文件
            with open(temp_file_path, 'wb') as temp_file:
                temp_file.write(content)

            # 验证数据库文件完整性
            conn = sqlite3.connect(temp_file_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()
            conn.close()

            # 检查是否包含必要的表
            table_names = [table[0] for table in tables]
            required_tables = ['users', 'cookies']  # 最基本的表

            missing_tables = [table for table in required_tables if table not in table_names]
            if missing_tables:
                log_with_user('warning', f"备份文件缺少必要的表: {missing_tables}", admin_user)
                raise HTTPException(status_code=400, detail=f"备份文件不完整，缺少表: {', '.join(missing_tables)}")

            log_with_user('info', f"备份文件验证通过，包含 {len(table_names)} 个表", admin_user)

        except sqlite3.Error as e:
            log_with_user('error', f"备份文件验证失败: {str(e)}", admin_user)
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            raise HTTPException(status_code=400, detail="无效的数据库文件")

        # 备份当前数据库
        from db_manager import db_manager
        current_db_path = db_manager.db_path

        # 生成备份文件路径（与原数据库在同一目录）
        db_dir = os.path.dirname(current_db_path)
        backup_filename = f"xianyu_data_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        backup_current_path = os.path.join(db_dir, backup_filename)

        if os.path.exists(current_db_path):
            shutil.copy2(current_db_path, backup_current_path)
            log_with_user('info', f"当前数据库已备份为: {backup_current_path}", admin_user)

        # 关闭当前数据库连接
        if hasattr(db_manager, 'conn') and db_manager.conn:
            db_manager.conn.close()
            log_with_user('info', "已关闭当前数据库连接", admin_user)

        # 替换数据库文件
        shutil.move(temp_file_path, current_db_path)
        log_with_user('info', f"数据库文件已替换: {current_db_path}", admin_user)

        # 重新初始化数据库连接（使用原有的db_path）
        db_manager.__init__(db_manager.db_path)
        log_with_user('info', "数据库连接已重新初始化", admin_user)

        # 验证新数据库
        try:
            test_users = db_manager.get_all_users()
            log_with_user('info', f"数据库恢复成功，包含 {len(test_users)} 个用户", admin_user)
        except Exception as e:
            log_with_user('error', f"数据库恢复后验证失败: {str(e)}", admin_user)
            # 如果验证失败，尝试恢复原数据库
            if os.path.exists(backup_current_path):
                shutil.copy2(backup_current_path, current_db_path)
                db_manager.__init__()
                log_with_user('info', "已恢复原数据库", admin_user)
            raise HTTPException(status_code=500, detail="数据库恢复失败，已回滚到原数据库")

        return {
            "success": True,
            "message": "数据库恢复成功",
            "backup_file": backup_current_path,
            "user_count": len(test_users)
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"上传数据库备份失败: {str(e)}", admin_user)
        # 清理临时文件
        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/backup/list')
def list_backup_files(admin_user: Dict[str, Any] = Depends(require_admin)):
    """列出服务器上的备份文件（管理员专用）"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', "查询备份文件列表", admin_user)

        # 查找备份文件（在data目录中）
        backup_files = glob.glob("data/xianyu_data_backup_*.db")

        backup_list = []
        for file_path in backup_files:
            try:
                stat = os.stat(file_path)
                backup_list.append({
                    'filename': os.path.basename(file_path),
                    'size': stat.st_size,
                    'size_mb': round(stat.st_size / (1024 * 1024), 2),
                    'created_time': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                    'modified_time': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
            except Exception as e:
                log_with_user('warning', f"读取备份文件信息失败: {file_path} - {str(e)}", admin_user)

        # 按修改时间倒序排列
        backup_list.sort(key=lambda x: x['modified_time'], reverse=True)

        log_with_user('info', f"找到 {len(backup_list)} 个备份文件", admin_user)

        return {
            "backups": backup_list,
            "total": len(backup_list)
        }

    except Exception as e:
        log_with_user('error', f"查询备份文件列表失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 系统管理接口 -------------------------

@app.post('/admin/reload-cache')
async def reload_system_cache(admin_user: Dict[str, Any] = Depends(require_admin)):
    """刷新系统缓存（管理员专用）"""
    try:
        log_with_user('info', "刷新系统缓存", admin_user)
        
        # 这里可以添加实际的缓存刷新逻辑
        # 例如：重新加载配置、清理内存缓存等
        
        log_with_user('info', "系统缓存刷新成功", admin_user)
        return {"success": True, "message": "系统缓存已刷新"}
        
    except Exception as e:
        log_with_user('error', f"刷新系统缓存失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 数据管理接口 -------------------------

@app.get('/admin/data/{table_name}')
def get_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取指定表的所有数据（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"查询表数据: {table_name}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', "item_replay",
            'risk_control_logs'
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试访问不允许的表: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许访问该表")

        # 获取表数据
        data, columns = db_manager.get_table_data(table_name)

        log_with_user('info', f"表 {table_name} 查询成功，共 {len(data)} 条记录", admin_user)

        return {
            "success": True,
            "data": data,
            "columns": columns,
            "count": len(data)
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"查询表数据失败: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/data/{table_name}/{record_id}')
def delete_table_record(table_name: str, record_id: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """删除指定表的指定记录（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"删除表记录: {table_name}.{record_id}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders','item_replay'
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试删除不允许的表记录: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许操作该表")

        # 特殊保护：不能删除管理员用户
        if table_name == 'users' and record_id == str(admin_user['user_id']):
            log_with_user('warning', "尝试删除管理员自己", admin_user)
            raise HTTPException(status_code=400, detail="不能删除管理员自己")

        # 删除记录
        success = db_manager.delete_table_record(table_name, record_id)

        if success:
            log_with_user('info', f"表记录删除成功: {table_name}.{record_id}", admin_user)
            return {"success": True, "message": "删除成功"}
        else:
            log_with_user('warning', f"表记录删除失败: {table_name}.{record_id}", admin_user)
            raise HTTPException(status_code=400, detail="删除失败，记录可能不存在")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除表记录异常: {table_name}.{record_id} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/data/{table_name}')
def clear_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """清空指定表的所有数据（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"清空表数据: {table_name}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', 'item_replay',
            'risk_control_logs'
        ]

        # 不允许清空用户表
        if table_name == 'users':
            log_with_user('warning', "尝试清空用户表", admin_user)
            raise HTTPException(status_code=400, detail="不允许清空用户表")

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试清空不允许的表: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许清空该表")

        # 清空表数据
        success = db_manager.clear_table_data(table_name)

        if success:
            log_with_user('info', f"表数据清空成功: {table_name}", admin_user)
            return {"success": True, "message": "清空成功"}
        else:
            log_with_user('warning', f"表数据清空失败: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="清空失败")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"清空表数据异常: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# 商品多规格管理API
@app.put("/items/{cookie_id}/{item_id}/multi-spec")
def update_item_multi_spec(cookie_id: str, item_id: str, spec_data: dict, _: None = Depends(require_auth)):
    """更新商品的多规格状态"""
    try:
        from db_manager import db_manager

        is_multi_spec = spec_data.get('is_multi_spec', False)

        success = db_manager.update_item_multi_spec_status(cookie_id, item_id, is_multi_spec)

        if success:
            return {"message": f"商品多规格状态已{'开启' if is_multi_spec else '关闭'}"}
        else:
            raise HTTPException(status_code=404, detail="商品不存在")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# 商品多数量发货管理API
@app.put("/items/{cookie_id}/{item_id}/multi-quantity-delivery")
def update_item_multi_quantity_delivery(cookie_id: str, item_id: str, delivery_data: dict, _: None = Depends(require_auth)):
    """更新商品的多数量发货状态"""
    try:
        from db_manager import db_manager

        multi_quantity_delivery = delivery_data.get('multi_quantity_delivery', False)

        success = db_manager.update_item_multi_quantity_delivery_status(cookie_id, item_id, multi_quantity_delivery)

        if success:
            return {"message": f"商品多数量发货状态已{'开启' if multi_quantity_delivery else '关闭'}"}
        else:
            raise HTTPException(status_code=404, detail="商品不存在")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





# ==================== 订单管理接口 ====================

@app.get('/api/orders')
def get_user_orders(
    current_user: Dict[str, Any] = Depends(get_current_user),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    cookie_id: Optional[str] = Query(None, description="筛选Cookie ID"),
    status: Optional[str] = Query(None, description="筛选状态")
):
    """获取当前用户的订单信息（支持分页）"""
    try:
        from db_manager import db_manager

        user_id = current_user['user_id']
        log_with_user('info', f"查询用户订单信息 (page={page}, page_size={page_size})", current_user)

        # 获取用户的所有Cookie
        user_cookies = db_manager.get_all_cookies(user_id)

        # 如果指定了cookie_id筛选
        if cookie_id and cookie_id in user_cookies:
            user_cookies = {cookie_id: user_cookies[cookie_id]}

        # 获取所有订单数据
        all_orders = []
        # 先获取所有商品的 item_id 到 item_title 的映射
        item_titles = {}
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            cursor.execute('SELECT item_id, item_title FROM item_info')
            for row in cursor.fetchall():
                item_titles[row[0]] = row[1]

        for cid in user_cookies.keys():
            orders = db_manager.get_orders_by_cookie(cid, limit=1000)
            for order in orders:
                order['cookie_id'] = cid
                # 添加 item_title 字段
                order['item_title'] = item_titles.get(order.get('item_id'), '')
                # 状态筛选
                if status and order.get('status') != status:
                    continue
                all_orders.append(order)

        # 按创建时间倒序排列
        all_orders.sort(key=lambda x: x.get('created_at', ''), reverse=True)

        # 分页处理
        total = len(all_orders)
        total_pages = (total + page_size - 1) // page_size
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_orders = all_orders[start_idx:end_idx]

        log_with_user('info', f"用户订单查询成功，共 {total} 条记录，第 {page}/{total_pages} 页", current_user)
        return {
            "success": True,
            "data": paginated_orders,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages
        }

    except Exception as e:
        log_with_user('error', f"查询用户订单失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"查询订单失败: {str(e)}")


@app.get('/api/orders/{order_id}')
def get_order_detail(order_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取订单详情"""
    try:
        from db_manager import db_manager

        user_id = current_user['user_id']
        log_with_user('info', f"查询订单详情: {order_id}", current_user)

        # 获取用户的所有Cookie
        user_cookies = db_manager.get_all_cookies(user_id)

        # 在用户的订单中查找
        for cookie_id in user_cookies.keys():
            order = db_manager.get_order_by_id(order_id)
            if order and order.get('cookie_id') == cookie_id:
                log_with_user('info', f"订单详情查询成功: {order_id}", current_user)
                return {"success": True, "data": order}

        log_with_user('warning', f"订单不存在或无权访问: {order_id}", current_user)
        raise HTTPException(status_code=404, detail="订单不存在或无权访问")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"查询订单详情失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"查询订单详情失败: {str(e)}")


@app.delete('/api/orders/{order_id}')
def delete_order(order_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除订单"""
    try:
        from db_manager import db_manager

        user_id = current_user['user_id']
        log_with_user('info', f"删除订单: {order_id}", current_user)

        # 获取用户的所有Cookie
        user_cookies = db_manager.get_all_cookies(user_id)

        # 验证订单属于当前用户
        order = db_manager.get_order_by_id(order_id)
        if not order:
            raise HTTPException(status_code=404, detail="订单不存在")

        if order.get('cookie_id') not in user_cookies:
            raise HTTPException(status_code=403, detail="无权删除此订单")

        # 删除订单
        success = db_manager.delete_order(order_id)
        if success:
            log_with_user('info', f"订单删除成功: {order_id}", current_user)
            return {"success": True, "message": "删除成功"}
        else:
            raise HTTPException(status_code=500, detail="删除失败")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除订单失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"删除订单失败: {str(e)}")


@app.post('/api/orders/{order_id}/refresh')
async def refresh_single_order(
    order_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """刷新单条订单状态"""
    try:
        from db_manager import db_manager
        from utils.order_fetcher_optimized import process_orders_batch

        user_id = current_user['user_id']
        log_with_user('info', f"刷新单条订单: {order_id}", current_user)

        # 获取用户的所有Cookie
        user_cookies = db_manager.get_all_cookies(user_id)

        # 验证订单存在且属于当前用户
        order = db_manager.get_order_by_id(order_id)
        if not order:
            raise HTTPException(status_code=404, detail="订单不存在")

        cookie_id = order.get('cookie_id')
        if not cookie_id or cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权刷新此订单")

        cookies_str = user_cookies[cookie_id]
        if not cookies_str:
            raise HTTPException(status_code=400, detail="Cookie无效")

        # 调用批量刷新函数处理单条订单
        batch_results = await process_orders_batch(
            order_ids=[order_id],
            cookie_id=cookie_id,
            cookie_string=cookies_str,
            max_concurrent=1,
            timeout=30,
            headless=True,
            use_pool=True,
            force_refresh=True
        )

        if not batch_results or len(batch_results) == 0:
            raise HTTPException(status_code=500, detail="刷新失败")

        result = batch_results[0]
        if result.get('error'):
            raise HTTPException(status_code=500, detail=f"刷新失败: {result.get('error')}")

        # 状态码映射
        order_status = result.get('order_status', 'unknown')
        if order_status and str(order_status).isdigit():
            status_mapping = {
                '1': 'processing',
                '2': 'pending_ship',
                '3': 'shipped',
                '4': 'completed',
                '5': 'refunding',
                '6': 'cancelled',
                '7': 'refunding',
                '8': 'cancelled',
                '9': 'refunding',
                '10': 'cancelled',
                '11': 'completed',
                '12': 'cancelled',
            }
            order_status = status_mapping.get(str(order_status), order_status)

        # 更新数据库
        db_manager.insert_or_update_order(
            order_id=order_id,
            item_id=result.get('item_id') or None,
            buyer_id=result.get('buyer_id') or None,
            spec_name=result.get('spec_name') or None,
            spec_value=result.get('spec_value') or None,
            quantity=result.get('quantity') or None,
            amount=result.get('amount') or None,
            order_status=order_status,
            cookie_id=cookie_id,
            receiver_name=result.get('receiver_name') or None,
            receiver_phone=result.get('receiver_phone') or None,
            receiver_address=result.get('receiver_address') or None,
        )

        log_with_user('info', f"订单刷新成功: {order_id}, 新状态: {order_status}", current_user)
        return JSONResponse({
            "success": True,
            "message": "订单刷新成功",
            "data": {
                "order_id": order_id,
                "order_status": order_status,
            }
        })

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"刷新订单失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"刷新订单失败: {str(e)}")


def check_order_data_completeness(order: Dict[str, Any]) -> bool:
    """
    检查订单数据是否完整

    Args:
        order: 订单数据字典

    Returns:
        True表示数据完整，False表示需要刷新
    """
    # 检查关键字段是否为空或为'unknown'
    incomplete_conditions = [
        not order.get('receiver_name') or order.get('receiver_name') == 'unknown',
        not order.get('receiver_phone') or order.get('receiver_phone') == 'unknown',
        not order.get('receiver_address') or order.get('receiver_address') == 'unknown',
        order.get('order_status') == 'unknown',
        not order.get('buyer_id') or order.get('buyer_id') == 'unknown',
    ]

    return not any(incomplete_conditions)


@app.put('/api/orders/{order_id}')
async def update_order(
    order_id: str,
    update_data: dict,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    更新订单信息
    自动检查订单数据完整性，如数据不完整则通过 Playwright 从订单详情页获取最新完整数据
    获取完整信息包括：订单ID、商品ID、买家ID、规格、数量、金额、订单状态、收货人信息
    """
    try:
        from db_manager import db_manager
        from utils.order_fetcher_optimized import fetch_order_complete

        user_id = current_user['user_id']
        log_with_user('info', f"更新订单: {order_id}, 数据: {update_data}", current_user)

        # 获取用户的所有Cookie
        user_cookies = db_manager.get_all_cookies(user_id)

        # 验证订单属于当前用户
        order = db_manager.get_order_by_id(order_id)
        if not order:
            raise HTTPException(status_code=404, detail="订单不存在")

        if order.get('cookie_id') not in user_cookies:
            raise HTTPException(status_code=403, detail="无权修改此订单")

        # 检查订单数据完整性
        is_complete = check_order_data_completeness(order)

        if not is_complete:
            log_with_user('info', f"订单 {order_id} 数据不完整，开始使用Playwright获取完整数据", current_user)

            # 获取该订单对应的Cookie字符串
            cookie_id = order.get('cookie_id')
            cookie_string = user_cookies.get(cookie_id)

            if cookie_string:

                try:
                    # 使用优化后的合并函数：一次浏览器访问获取所有数据
                    log_with_user('info', f"使用优化方法获取订单 {order_id} 的完整数据", current_user)

                    complete_result = await fetch_order_complete(
                        order_id=order_id,
                        cookie_id=cookie_id,
                        cookie_string=cookie_string,
                        timeout=30,
                        headless=True,
                        use_pool=True  # 使用浏览器池
                    )

                    if complete_result:
                        log_with_user('info', f"成功获取订单 {order_id} 的完整数据（一次浏览器调用）", current_user)

                        # 状态码映射（如果需要转换）
                        order_status = complete_result.get('order_status', 'unknown')
                        if order_status and isinstance(order_status, str) and order_status.isdigit():
                            status_mapping = {
                                '1': 'processing',
                                '2': 'pending_ship',
                                '3': 'shipped',
                                '4': 'completed',
                                '5': 'refunding',
                                '6': 'cancelled',
                                '7': 'refunding',
                                '8': 'cancelled',
                                '9': 'refunding',
                                '10': 'cancelled',
                            }
                            order_status = status_mapping.get(order_status, order_status)

                        # 构建要更新的完整数据
                        refresh_data = {
                            'order_id': order_id,
                            'item_id': complete_result.get('item_id') or order.get('item_id'),
                            'buyer_id': complete_result.get('buyer_id') or order.get('buyer_id'),
                            'order_status': order_status or order.get('order_status'),
                            'spec_name': complete_result.get('spec_name') or None,
                            'spec_value': complete_result.get('spec_value') or None,
                            'quantity': complete_result.get('quantity') or None,
                            'amount': complete_result.get('amount') or None,
                            'created_at': complete_result.get('order_time') or None,
                            'receiver_name': complete_result.get('receiver_name') or None,
                            'receiver_phone': complete_result.get('receiver_phone') or None,
                            'receiver_address': complete_result.get('receiver_address') or None
                        }

                        # 更新数据库
                        db_manager.insert_or_update_order(**refresh_data)
                        log_with_user('info', f"订单 {order_id} 完整数据已更新到数据库", current_user)
                    else:
                        log_with_user('warning', f"订单 {order_id} 详情获取失败，继续使用现有数据", current_user)

                except Exception as e:
                    log_with_user('error', f"获取订单 {order_id} 详情时出错: {str(e)}", current_user)
                    # 继续执行，即使刷新失败也允许用户手动更新
            else:
                log_with_user('warning', f"订单 {order_id} 的Cookie信息不完整，无法刷新", current_user)

        # 提取可更新的字段
        allowed_fields = {
            'item_id', 'buyer_id', 'spec_name', 'spec_value',
            'quantity', 'amount', 'order_status',
            'receiver_name', 'receiver_phone', 'receiver_address',
            'system_shipped', 'created_at'
        }

        # 只保留允许更新的字段
        filtered_data = {k: v for k, v in update_data.items() if k in allowed_fields}

        if not filtered_data:
            # 如果没有用户提供的更新数据
            if not is_complete:
                # 数据不完整，已经进行了自动刷新，返回刷新后的订单
                updated_order = db_manager.get_order_by_id(order_id)
                return {
                    "success": True,
                    "message": "订单数据已自动刷新",
                    "data": updated_order,
                    "refreshed": True
                }
            else:
                # 数据完整，直接返回当前订单信息
                updated_order = db_manager.get_order_by_id(order_id)
                return {
                    "success": True,
                    "message": "订单数据已是最新",
                    "data": updated_order,
                    "refreshed": False
                }

        # 应用用户提供的更新
        success = db_manager.insert_or_update_order(
            order_id=order_id,
            **filtered_data
        )

        if success:
            log_with_user('info', f"订单更新成功: {order_id}", current_user)
            # 返回更新后的订单
            updated_order = db_manager.get_order_by_id(order_id)
            return {
                "success": True,
                "message": "更新成功",
                "data": updated_order,
                "refreshed": not is_complete  # 标记是否进行了自动刷新
            }
        else:
            raise HTTPException(status_code=500, detail="更新失败")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"更新订单失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"更新订单失败: {str(e)}")


@app.post('/api/orders/refresh')
async def refresh_orders_status(
    cookie_id: Optional[str] = Form(None),
    status: Optional[str] = Form(None),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    智能刷新订单状态
    1. 从数据库获取订单列表（支持筛选）
    2. 对非'已发货'状态的订单，使用Playwright查询最新状态
    3. 更新数据库中有变化的订单
    """
    try:
        from db_manager import db_manager
        from utils.order_fetcher_optimized import process_orders_batch

        user_id = current_user['user_id']
        log_with_user('info', f"开始智能刷新订单状态（优化版：并发处理） (cookie_id={cookie_id}, status={status})", current_user)

        # 获取用户的所有Cookie
        user_cookies = db_manager.get_all_cookies(user_id)

        # 如果指定了cookie_id，只使用该Cookie
        if cookie_id:
            if cookie_id not in user_cookies:
                raise HTTPException(status_code=404, detail="Cookie不存在或无权访问")
            user_cookies = {cookie_id: user_cookies[cookie_id]}

        # 获取需要刷新的订单
        orders_to_refresh = []
        for cid in user_cookies.keys():
            # 获取该Cookie的所有订单
            orders = db_manager.get_orders_by_cookie(cid, limit=1000)

            # 筛选需要刷新的订单
            for order in orders:
                # 如果指定了状态筛选，只刷新该状态的订单
                if status and order.get('status') != status:
                    continue

                order_status = order.get('status', 'unknown')

                # 判断是否需要刷新：只根据状态判断
                # 稳定状态（已发货、交易成功、交易关闭）的订单不需要刷新
                needs_refresh = order_status not in ['shipped', 'completed', 'cancelled']

                if needs_refresh:
                    orders_to_refresh.append({
                        'order_id': order['order_id'],
                        'cookie_id': cid,
                        'current_status': order_status
                    })

        log_with_user('info', f"找到 {len(orders_to_refresh)} 个需要刷新的订单", current_user)

        if not orders_to_refresh:
            return JSONResponse({
                "success": True,
                "message": "没有需要刷新的订单",
                "summary": {
                    "total": 0,
                    "updated": 0,
                    "no_change": 0,
                    "failed": 0
                },
                "results": []
            })

        # 刷新订单信息（包括状态、买家ID、金额等）
        updated_count = 0
        failed_count = 0
        no_change_count = 0
        refresh_results = []

        # 按cookie_id分组订单（因为每个cookie需要单独的浏览器实例）
        orders_by_cookie = {}
        for order_info in orders_to_refresh:
            cid = order_info['cookie_id']
            if cid not in orders_by_cookie:
                orders_by_cookie[cid] = []
            orders_by_cookie[cid].append(order_info)

        # 对每个cookie的订单进行并发批量处理
        for cid, cookie_orders in orders_by_cookie.items():
            cookies_str = user_cookies[cid]
            if not cookies_str:
                log_with_user('warning', f"Cookie {cid} 的值为空，跳过", current_user)
                failed_count += len(cookie_orders)
                continue

            # 提取订单ID列表
            order_ids = [o['order_id'] for o in cookie_orders]
            log_with_user('info', f"使用并发处理Cookie {cid} 的 {len(order_ids)} 个订单", current_user)

            # 并发批量处理（一次浏览器调用获取所有数据）
            batch_results = await process_orders_batch(
                order_ids=order_ids,
                cookie_id=cid,
                cookie_string=cookies_str,
                max_concurrent=5,  # 并发5个
                timeout=30,
                headless=True,
                use_pool=True,  # 使用浏览器池
                force_refresh=True  # 强制刷新，跳过缓存检查
            )

            # 处理结果并更新数据库
            for i, result in enumerate(batch_results):
                order_info = cookie_orders[i]
                order_id = order_info['order_id']
                current_status = order_info['current_status']

                if result and not result.get('error'):
                    # 调试：打印API和DOM状态
                    api_status = result.get('api_status', 'N/A')
                    dom_status = result.get('dom_status', 'N/A')
                    log_with_user('debug', f"订单 {order_id} - API状态: {api_status}, DOM状态: {dom_status}", current_user)

                    # 状态码映射
                    order_status = result.get('order_status', 'unknown')
                    if order_status and str(order_status).isdigit():
                        status_mapping = {
                            '1': 'processing',
                            '2': 'pending_ship',
                            '3': 'shipped',
                            '4': 'completed',
                            '5': 'refunding',
                            '6': 'cancelled',
                            '7': 'refunding',
                            '8': 'cancelled',
                            '9': 'refunding',
                            '10': 'cancelled',
                            '11': 'completed',  # 交易完成
                            '12': 'cancelled',  # 交易关闭
                        }
                        order_status = status_mapping.get(str(order_status), order_status)

                    # 更新数据库
                    success = db_manager.insert_or_update_order(
                        order_id=order_id,
                        item_id=result.get('item_id') or None,
                        buyer_id=result.get('buyer_id') or None,
                        spec_name=result.get('spec_name') or None,
                        spec_value=result.get('spec_value') or None,
                        quantity=result.get('quantity') or None,
                        amount=result.get('amount') or None,
                        order_status=order_status if order_status != current_status else None,
                        cookie_id=cid,
                        created_at=result.get('order_time') or None,
                        receiver_name=result.get('receiver_name') or None,
                        receiver_phone=result.get('receiver_phone') or None,
                        receiver_address=result.get('receiver_address') or None
                    )

                    if success:
                        # 检查是否有更新
                        has_changes = (
                            order_status != current_status or
                            result.get('buyer_id') or
                            result.get('amount')
                        )

                        if has_changes:
                            updated_count += 1
                            refresh_results.append({
                                'order_id': order_id,
                                'old_status': current_status,
                                'new_status': order_status,
                                'status_text': result.get('status_text', '')
                            })
                            log_with_user('info', f"订单 {order_id} 已更新 | {current_status} -> {order_status}", current_user)
                        else:
                            no_change_count += 1
                    else:
                        failed_count += 1
                        log_with_user('error', f"订单 {order_id} 更新失败", current_user)
                else:
                    failed_count += 1
                    error_msg = result.get('error', '未知错误') if result else '未知错误'
                    log_with_user('warning', f"订单 {order_id} 获取失败: {error_msg}", current_user)

        # 由于我们已经处理完所有订单，跳过原来的循环
        # 下面的代码需要删除，所以我们需要找到循环结束的位置
        if False:  # 这个if永远不会执行，只是为了保持代码结构
            for order_info in orders_to_refresh:
                order_id = order_info['order_id']
                cookie_id = order_info['cookie_id']
                current_status = order_info['current_status']

                try:
                    # 获取Cookie (get_all_cookies返回的是 {cookie_id: cookie_value} 格式)
                    cookies_str = user_cookies[cookie_id]

                    if not cookies_str:
                        log_with_user('warning', f"Cookie {cookie_id} 的值为空，跳过订单 {order_id}", current_user)
                        failed_count += 1
                        continue

                    # 使用订单详情获取器获取完整信息（包括买家ID、金额、收货人信息）
                    # 注意：fetch_order_detail_simple 已经能获取所有需要的数据，无需再调用 OrderStatusQueryPlaywright
                    order_detail = await fetch_order_detail_simple(order_id, cookies_str, headless=True)

                    if order_detail:
                        # 提取订单详情（从页面获取）
                        spec_name = order_detail.get('spec_name', '')
                        spec_value = order_detail.get('spec_value', '')
                        quantity = order_detail.get('quantity', '')
                        amount = order_detail.get('amount', '')
                        receiver_name = order_detail.get('receiver_name', '')
                        receiver_phone = order_detail.get('receiver_phone', '')
                        receiver_address = order_detail.get('receiver_address', '')

                        # 只使用状态查询获取订单状态和买家ID（因为DOM解析无法获取这些）
                        query = OrderStatusQueryPlaywright(cookies_str, cookie_id, headless=True)
                        status_result = await query.query_order_status(order_id)

                        new_status = current_status
                        new_status_text = ''
                        buyer_id = ''
                        item_id = ''
                        is_bargain = None

                        if status_result.get('success'):
                            new_status_code = status_result.get('order_status')
                            new_status_text = status_result.get('status_text', '')

                            # 将状态码转换为数据库状态
                            # 完整的订单状态码映射（基于闲鱼API）
                            status_mapping = {
                                1: 'processing',      # 处理中
                                2: 'pending_ship',    # 待发货
                                3: 'shipped',         # 已发货
                                4: 'completed',       # 已完成/交易成功
                                5: 'refunding',       # 退款中
                                6: 'cancelled',       # 已取消/已关闭
                                7: 'refunding',       # 退款申请中
                                8: 'cancelled',       # 退款成功（订单关闭）
                                9: 'refunding',       # 退款协商中
                                10: 'cancelled',      # 退款关闭
                            }
                            new_status = status_mapping.get(new_status_code, 'unknown')

                            # 特殊处理：根据状态文本智能识别（优先检查最终状态）
                            if new_status == 'unknown':
                                # 优先级1: 检查"退款成功"（最终状态）
                                if '退款' in new_status_text and '成功' in new_status_text:
                                    new_status = 'cancelled'  # 退款成功=订单关闭
                                # 优先级2: 检查"关闭"或"取消"（最终状态）
                                elif '关闭' in new_status_text or '取消' in new_status_text or '超时' in new_status_text:
                                    new_status = 'cancelled'
                                # 优先级3: 检查"完成"或"交易成功"（最终状态）
                                elif '完成' in new_status_text or '交易成功' in new_status_text or '确认收货' in new_status_text:
                                    new_status = 'completed'
                                # 优先级4: 检查"退款"（中间状态）
                                elif '退款' in new_status_text:
                                    new_status = 'refunding'

                            log_with_user('debug', f"订单 {order_id}: 状态码={new_status_code}, 状态文本={new_status_text}, 映射结果={new_status}", current_user)

                            # 从 raw_data 中提取完整信息
                            raw_data = status_result.get('raw_data', {})

                            # 提取买家ID、商品ID、时间信息
                            created_at = None
                            try:
                                # 方法1: 从根级别提取 peerUserId (买家ID)
                                buyer_id = str(raw_data.get('peerUserId', ''))

                                # 方法2: 从根级别提取 itemId (商品ID)
                                item_id = str(raw_data.get('itemId', ''))

                                # 方法3: 从 orderStatusVO 组件中提取下单时间
                                if 'components' in raw_data:
                                    for component in raw_data['components']:
                                        if component.get('render') == 'orderStatusVO':
                                            order_status_data = component.get('data', {})
                                            # 从 orderStatusNodeList 中找到第一个时间节点（已拍下时间 = 创建时间）
                                            node_list = order_status_data.get('orderStatusNodeList', [])
                                            if node_list and len(node_list) > 0:
                                                created_at = node_list[0].get('time')  # 第一个是"已拍下"时间
                                            break

                                # 方法4: 从 orderInfoVO 组件中提取是否小刀（如果有 bargainInfo）
                                if 'components' in raw_data:
                                    for component in raw_data['components']:
                                        if component.get('render') == 'orderInfoVO':
                                            data = component.get('data', {})
                                            # 检查是否有小刀信息
                                            if 'bargainInfo' in data:
                                                bargain_info = data.get('bargainInfo', {})
                                                is_bargain = bargain_info.get('bargain', False)
                                            # 如果前面没找到商品ID，尝试从 jumpUrl 中提取
                                            if not item_id:
                                                item_info = data.get('itemInfo', {})
                                                jump_url = item_info.get('jumpUrl', '')
                                                if 'id=' in jump_url:
                                                    item_id = jump_url.split('id=')[1].split('&')[0]
                                            break

                                if created_at:
                                    log_with_user('debug', f"提取到订单创建时间: {created_at}", current_user)

                            except Exception as e:
                                log_with_user('warning', f"提取订单信息失败: {str(e)}", current_user)

                        # 更新数据库（包含所有字段）
                        success = db_manager.insert_or_update_order(
                            order_id=order_id,
                            item_id=item_id if item_id else None,
                            buyer_id=buyer_id if buyer_id else None,
                            spec_name=spec_name if spec_name else None,
                            spec_value=spec_value if spec_value else None,
                            quantity=quantity if quantity else None,
                            amount=amount if amount else None,
                            order_status=new_status if new_status != current_status else None,
                            is_bargain=is_bargain if is_bargain is not None else None,
                            cookie_id=cookie_id,
                            created_at=created_at,  # 添加创建时间（从API提取的北京时间）
                            receiver_name=receiver_name if receiver_name else None,
                            receiver_phone=receiver_phone if receiver_phone else None,
                            receiver_address=receiver_address if receiver_address else None
                        )

                        if success:
                            # 检查是否有任何更新
                            has_changes = (
                                new_status != current_status or
                                (buyer_id and buyer_id != 'unknown_user') or
                                amount
                            )

                            if has_changes:
                                updated_count += 1
                                refresh_results.append({
                                    'order_id': order_id,
                                    'old_status': current_status,
                                    'new_status': new_status,
                                    'status_text': new_status_text
                                })
                                log_with_user('info', f"订单 {order_id} 信息已更新 | 状态: {current_status} -> {new_status} | 买家: {buyer_id} | 金额: {amount}", current_user)
                            else:
                                no_change_count += 1
                                log_with_user('debug', f"订单 {order_id} 信息无变化", current_user)
                        else:
                            failed_count += 1
                            log_with_user('error', f"订单 {order_id} 信息更新失败", current_user)
                    else:
                        failed_count += 1
                        log_with_user('warning', f"订单 {order_id} 详情获取失败", current_user)

                except Exception as e:
                    failed_count += 1
                    log_with_user('error', f"刷新订单 {order_id} 时发生异常: {str(e)}", current_user)

        # 返回刷新结果
        log_with_user('info', f"订单刷新完成: 更新{updated_count}个, 无变化{no_change_count}个, 失败{failed_count}个", current_user)

        return JSONResponse({
            "success": True,
            "message": f"刷新完成: 更新{updated_count}个, 无变化{no_change_count}个, 失败{failed_count}个",
            "summary": {
                "total": len(orders_to_refresh),
                "updated": updated_count,
                "no_change": no_change_count,
                "failed": failed_count
            },
            "updated_orders": refresh_results
        })

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"刷新订单状态失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"刷新订单状态失败: {str(e)}")


# 已取消：全量核对订单数据功能
# 现在使用更新订单状态接口进行单个订单的数据核查
# @app.post('/api/orders/verify-all')
# async def verify_all_orders(current_user: Dict[str, Any] = Depends(get_current_user)):
#     """
#     全量核对所有订单数据
#     通过 Playwright 访问每个订单的详情页，更新时间、收货人信息等
#     """
#     pass


@app.post('/api/orders/manual-ship')
async def manual_ship_orders(
    order_ids: List[str] = Body(..., description="订单ID列表"),
    ship_mode: str = Body(..., description="发货模式: status_only（仅修改发货状态）或 full_delivery（完整发货流程）"),
    custom_content: Optional[str] = Body(None, description="自定义发货内容（保留兼容）"),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    手动发货

    发货模式：
    - status_only: 仅在闲鱼标记为已发货（不发送卡券给买家）
    - full_delivery: 完整发货流程（匹配卡券、发送卡券给买家、标记发货状态）
    """
    try:
        from db_manager import db_manager
        from XianyuAutoAsync import XianyuLive
        import asyncio

        user_id = current_user['user_id']
        log_with_user('info', f"开始手动发货: 订单数量={len(order_ids)}, 模式={ship_mode}", current_user)

        # 验证发货模式
        if ship_mode not in ['status_only', 'full_delivery']:
            raise HTTPException(status_code=400, detail="发货模式必须是 status_only 或 full_delivery")

        # 获取用户的所有Cookie
        user_cookies = db_manager.get_all_cookies(user_id)

        success_count = 0
        failed_count = 0
        results = []

        # 遍历每个订单
        for order_id in order_ids:
            try:
                # 获取订单信息
                order = db_manager.get_order_by_id(order_id)
                if not order:
                    results.append({
                        'order_id': order_id,
                        'success': False,
                        'message': '订单不存在'
                    })
                    failed_count += 1
                    continue

                # 验证订单属于当前用户
                cookie_id = order.get('cookie_id')
                if cookie_id not in user_cookies:
                    results.append({
                        'order_id': order_id,
                        'success': False,
                        'message': '无权操作此订单'
                    })
                    failed_count += 1
                    continue

                item_id = order.get('item_id')
                buyer_id = order.get('buyer_id')

                if ship_mode == 'status_only':
                    # ====== 仅修改闲鱼发货状态 ======
                    if not item_id:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '订单缺少商品ID'
                        })
                        failed_count += 1
                        continue

                    # 获取cookies_str用于创建独立session
                    cookies_str = user_cookies.get(cookie_id)
                    if not cookies_str:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '无法获取账号Cookie信息'
                        })
                        failed_count += 1
                        continue

                    # 创建独立的aiohttp session（避免跨异步上下文问题）
                    import aiohttp
                    from secure_confirm_decrypted import SecureConfirm

                    try:
                        async with aiohttp.ClientSession(
                            headers={'cookie': cookies_str},
                            timeout=aiohttp.ClientTimeout(total=30)
                        ) as session:
                            confirm = SecureConfirm(session, cookies_str, cookie_id, None)
                            confirm_result = await confirm.auto_confirm(order_id, item_id)

                        if confirm_result and confirm_result.get('success'):
                            # 更新本地数据库状态
                            db_manager.insert_or_update_order(
                                order_id=order_id,
                                order_status='shipped',
                                system_shipped=True
                            )
                            results.append({
                                'order_id': order_id,
                                'success': True,
                                'message': '已成功修改闲鱼发货状态'
                            })
                            success_count += 1
                        else:
                            error_msg = confirm_result.get('error', '未知错误') if confirm_result else '确认发货返回空结果'
                            results.append({
                                'order_id': order_id,
                                'success': False,
                                'message': f'修改发货状态失败: {error_msg}'
                            })
                            failed_count += 1
                    except Exception as e:
                        log_with_user('error', f"确认发货异常: {str(e)}", current_user)
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': f'确认发货异常: {str(e)}'
                        })
                        failed_count += 1

                elif ship_mode == 'full_delivery':
                    # ====== 完整发货流程：匹配卡券 + 发送卡券 + 修改状态 ======
                    if not item_id:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '订单缺少商品ID，无法匹配发货规则'
                        })
                        failed_count += 1
                        continue

                    if not buyer_id:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '订单缺少买家ID，无法发送卡券'
                        })
                        failed_count += 1
                        continue

                    # 必须有运行中的实例（需要WebSocket发送消息）
                    live_instance = XianyuLive.get_instance(cookie_id)
                    if not live_instance:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '该账号未在线运行，无法执行完整发货。请先启动账号。'
                        })
                        failed_count += 1
                        continue

                    if not live_instance.ws or live_instance.ws.closed:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '该账号WebSocket连接已断开，无法发送消息。请等待重连后重试。'
                        })
                        failed_count += 1
                        continue

                    # 查找与买家的chat_id（优先从订单记录获取，回退到AI对话记录）
                    chat_id = order.get('chat_id') or ''
                    if not chat_id:
                        chat_id = db_manager.find_chat_id_by_buyer(cookie_id, buyer_id)
                    if not chat_id:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '未找到与该买家的聊天记录，无法发送卡券消息。请等待买家发送消息后重试。'
                        })
                        failed_count += 1
                        continue

                    # 检查多数量发货
                    quantity_to_send = 1
                    multi_quantity_delivery = db_manager.get_item_multi_quantity_delivery_status(cookie_id, item_id)
                    if multi_quantity_delivery:
                        try:
                            order_detail = await live_instance.fetch_order_detail_info(order_id, item_id, buyer_id)
                            if order_detail and isinstance(order_detail, dict):
                                qty = order_detail.get('quantity', 1)
                                if isinstance(qty, int) and qty > 1:
                                    quantity_to_send = qty
                        except Exception as e:
                            log_with_user('warning', f"获取订单数量失败，使用默认数量1: {str(e)}", current_user)

                    # 调用_auto_delivery获取卡券内容（内部会调用auto_confirm）
                    delivery_contents = []
                    for i in range(quantity_to_send):
                        try:
                            delivery_content = await live_instance._auto_delivery(
                                item_id, '', order_id, buyer_id
                            )
                            if delivery_content:
                                delivery_contents.append(delivery_content)
                        except Exception as e:
                            log_with_user('error', f"获取第{i+1}个卡券失败: {str(e)}", current_user)

                    if not delivery_contents:
                        results.append({
                            'order_id': order_id,
                            'success': False,
                            'message': '未匹配到发货规则或卡券获取失败'
                        })
                        failed_count += 1
                        continue

                    # 发送卡券内容给买家
                    send_success = True
                    for idx, content in enumerate(delivery_contents):
                        try:
                            if content.startswith("__IMAGE_SEND__"):
                                image_data = content.replace("__IMAGE_SEND__", "")
                                card_id = None
                                if "|" in image_data:
                                    card_id_str, image_url = image_data.split("|", 1)
                                    try:
                                        card_id = int(card_id_str)
                                    except ValueError:
                                        card_id = None
                                else:
                                    image_url = image_data
                                await live_instance.send_image_msg(
                                    live_instance.ws, chat_id, buyer_id,
                                    image_url, card_id=card_id
                                )
                            else:
                                await live_instance.send_msg(
                                    live_instance.ws, chat_id, buyer_id, content
                                )

                            # 多条消息之间间隔1秒
                            if len(delivery_contents) > 1 and idx < len(delivery_contents) - 1:
                                await asyncio.sleep(1)
                        except Exception as e:
                            log_with_user('error', f"发送第{idx+1}条卡券消息失败: {str(e)}", current_user)
                            send_success = False

                    # 更新本地数据库状态
                    db_manager.insert_or_update_order(
                        order_id=order_id,
                        order_status='shipped',
                        system_shipped=True
                    )

                    if send_success:
                        results.append({
                            'order_id': order_id,
                            'success': True,
                            'message': f'完整发货成功，已发送{len(delivery_contents)}条卡券信息给买家'
                        })
                        success_count += 1
                    else:
                        results.append({
                            'order_id': order_id,
                            'success': True,
                            'message': f'发货状态已更新，但部分卡券消息发送失败（共{len(delivery_contents)}条）'
                        })
                        success_count += 1

            except Exception as e:
                results.append({
                    'order_id': order_id,
                    'success': False,
                    'message': str(e)
                })
                failed_count += 1
                log_with_user('error', f"发货订单 {order_id} 时发生异常: {str(e)}", current_user)

        log_with_user('info', f"手动发货完成: 成功{success_count}个, 失败{failed_count}个", current_user)

        return {
            "success": True,
            "message": f"发货完成: 成功{success_count}个, 失败{failed_count}个",
            "total": len(order_ids),
            "success_count": success_count,
            "failed_count": failed_count,
            "results": results
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"手动发货失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"手动发货失败: {str(e)}")


@app.post('/api/orders/import')
async def import_orders(
    orders: List[Dict[str, Any]] = Body(..., description="订单列表"),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    导入订单
    支持批量导入自定义订单数据
    """
    try:
        from db_manager import db_manager

        user_id = current_user['user_id']
        log_with_user('info', f"开始导入订单: 订单数量={len(orders)}", current_user)

        # 获取用户的所有Cookie
        user_cookies = db_manager.get_all_cookies(user_id)

        success_count = 0
        failed_count = 0
        results = []

        # 必需字段验证
        required_fields = ['order_id', 'cookie_id']
        optional_fields = [
            'item_id', 'item_title', 'item_price', 'item_image',
            'buyer_id',
            'receiver_name', 'receiver_phone', 'receiver_address', 'receiver_city',
            'status', 'status_text', 'order_time', 'pay_time',
            'quantity', 'amount'
        ]

        for order_data in orders:
            try:
                # 验证必需字段
                missing_fields = [f for f in required_fields if not order_data.get(f)]
                if missing_fields:
                    results.append({
                        'order_id': order_data.get('order_id', 'unknown'),
                        'success': False,
                        'message': f'缺少必需字段: {", ".join(missing_fields)}'
                    })
                    failed_count += 1
                    continue

                order_id = str(order_data['order_id'])
                cookie_id = str(order_data['cookie_id'])

                # 验证Cookie属于当前用户
                if cookie_id not in user_cookies:
                    results.append({
                        'order_id': order_id,
                        'success': False,
                        'message': '无权操作此账号的订单'
                    })
                    failed_count += 1
                    continue

                # 检查订单是否已存在
                existing_order = db_manager.get_order_by_id(order_id)

                # 准备订单数据，直接使用 insert_or_update_order 的参数名
                # 构建参数字典，只传递非 None 的值
                insert_params = {
                    'order_id': order_id,
                    'cookie_id': cookie_id
                }

                # 前端字段名 -> 数据库参数名映射
                param_mapping = {
                    'item_id': 'item_id',
                    'buyer_id': 'buyer_id',
                    'receiver_name': 'receiver_name',
                    'receiver_phone': 'receiver_phone',
                    'receiver_address': 'receiver_address',
                    'receiver_city': 'receiver_city',
                    'status': 'order_status',  # 注意：前端用 status，后端用 order_status
                    'status_text': 'status_text',
                    'order_time': 'order_time',
                    'pay_time': 'pay_time',
                    'quantity': 'quantity',
                    'amount': 'amount',
                    'item_title': 'item_title',
                    'item_price': 'item_price',
                    'item_image': 'item_image'
                }

                # 遍历订单数据，添加到参数字典
                for field, value in order_data.items():
                    if value is not None and field in param_mapping:
                        param_name = param_mapping[field]
                        insert_params[param_name] = value

                # 使用 insert_or_update_order 统一处理
                db_manager.insert_or_update_order(**insert_params)

                results.append({
                    'order_id': order_id,
                    'success': True,
                    'message': '订单已更新' if existing_order else '订单已导入'
                })

                success_count += 1

            except Exception as e:
                results.append({
                    'order_id': order_data.get('order_id', 'unknown'),
                    'success': False,
                    'message': str(e)
                })
                failed_count += 1
                log_with_user('error', f"导入订单时发生异常: {str(e)}", current_user)

        log_with_user('info', f"导入订单完成: 成功{success_count}个, 失败{failed_count}个", current_user)

        return {
            "success": True,
            "message": f"导入完成: 成功{success_count}个, 失败{failed_count}个",
            "total": len(orders),
            "success_count": success_count,
            "failed_count": failed_count,
            "results": results
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"导入订单失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"导入订单失败: {str(e)}")


# ==================== 前端 SPA Catch-All 路由 ====================
# 必须放在所有 API 路由之后，用于处理前端 SPA 的直接访问
# 这样用户直接访问 /dashboard、/accounts 等前端路由时，会返回 index.html
# 然后由 React Router 在客户端处理路由

# 定义不需要返回前端页面的路径前缀（API 路径）
API_PREFIXES = ['/api/', '/static/', '/assets', '/health', '/login', '/logout', '/verify', '/change-password', '/change-admin-password']

@app.get('/{path:path}', response_class=HTMLResponse)
async def catch_all_route(path: str):
    """
    Catch-all 路由：处理所有未匹配的 GET 请求
    如果是 API 请求，返回 404；否则返回前端 index.html
    """
    # 检查是否是 API 请求
    full_path = f'/{path}'
    for prefix in API_PREFIXES:
        if full_path.startswith(prefix):
            raise HTTPException(status_code=404, detail="Not Found")
    
    # 返回前端页面
    return await serve_frontend()


# 移除自动启动，由Start.py或手动启动
# if __name__ == "__main__":
#     uvicorn.run(app, host="0.0.0.0", port=8080)